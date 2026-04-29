"""
Polyspace Analysis Agent
Main entry point for Polyspace report analysis

Developer: Himanshu R
Organization: Qorix India Pvt Ltd
Version: 3.0.0
"""

import sys
import logging
import argparse
import json
from pathlib import Path
from datetime import datetime

# Configure UTF-8 encoding for Windows console
if sys.platform == 'win32':
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleOutputCP(65001)  # UTF-8
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except:
        pass

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))

from PolyspaceParser import PolyspaceParser
from KnowledgeDatabaseManager import KnowledgeDatabaseManager

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('polyspace_analysis.log', mode='a')
    ]
)
logger = logging.getLogger(__name__)


def main():
    """Main entry point for Polyspace analysis"""
    
    parser = argparse.ArgumentParser(
        description='Polyspace Report Analysis Tool v3.0.0',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python run_polyspace_agent.py report.html Mka
  python run_polyspace_agent.py report.xlsx Eep --source-code D:\\Code\\Eep
  python run_polyspace_agent.py results.tsv Can --no-fixes
        '''
    )
    
    parser.add_argument(
        'report_file',
        type=str,
        help='Path to Polyspace report file (HTML, Excel, or TSV)'
    )
    
    parser.add_argument(
        'module_name',
        type=str,
        help='Module name (e.g., Mka, Eep, Can)'
    )
    
    parser.add_argument(
        '--source-code',
        type=str,
        help='Path to source code directory (for context-aware fixes)',
        default=None
    )
    
    parser.add_argument(
        '--no-fixes',
        action='store_true',
        help='Skip automatic fix generation (only parse and create KB)'
    )
    
    parser.add_argument(
        '--output-dir',
        type=str,
        help='Output directory for reports (default: project reports folder)',
        default=None
    )
    
    args = parser.parse_args()
    
    # Setup paths
    project_root = Path(__file__).parent.parent
    reports_dir = Path(args.output_dir) if args.output_dir else project_root / 'reports'
    kb_dir = project_root / 'knowledge_base'
    data_dir = project_root / 'data'
    
    # Create directories
    reports_dir.mkdir(exist_ok=True)
    kb_dir.mkdir(exist_ok=True)
    data_dir.mkdir(exist_ok=True)
    
    logger.info("="*80)
    logger.info("Polyspace Analysis Tool v3.0.0")
    logger.info("Qorix India Pvt Ltd")
    logger.info("="*80)
    logger.info(f"Module: {args.module_name}")
    logger.info(f"Report: {args.report_file}")
    logger.info(f"Source code: {args.source_code or 'Not provided'}")
    logger.info("="*80)
    
    try:
        # Validate report file path
        report_path = Path(args.report_file)
        if not report_path.exists():
            logger.error(f"Report file not found: {report_path}")
            logger.error("Please provide a valid file path")
            return 1
        
        if report_path.is_dir():
            logger.error(f"ERROR: You provided a directory path, not a file path!")
            logger.error(f"Directory: {report_path}")
            logger.error("")
            logger.error("Please provide the full path to the Polyspace report FILE.")
            logger.error(f"Example: {report_path}\\polyspace_results.html")
            return 1
        
        # Step 1: Parse Polyspace report
        logger.info("STEP 1: Parsing Polyspace report...")
        parser_obj = PolyspaceParser()
        violations = parser_obj.parse_report(report_path)
        
        if not violations:
            logger.warning("No violations found in report")
            return 0
        
        logger.info(f"Found {len(violations)} Polyspace findings")
        
        # Display summary
        summary = parser_obj.get_summary()
        logger.info("="*80)
        logger.info("Polyspace Findings Summary:")
        logger.info(f"  Total: {summary['total_findings']}")
        logger.info(f"  🔴 Red (Proven errors): {summary['red_findings']}")
        logger.info(f"  🟠 Orange (Potential errors): {summary['orange_findings']}")
        logger.info(f"  🟢 Green (Proven safe): {summary['green_findings']}")
        logger.info(f"  ⚪ Gray (Unreachable): {summary['gray_findings']}")
        logger.info("="*80)
        
        # Step 2: Save violations to JSON
        logger.info("STEP 2: Saving violations data...")
        violations_file = data_dir / f"{args.module_name}_violations.json"
        with open(violations_file, 'w', encoding='utf-8') as f:
            json.dump({
                'module_name': args.module_name,
                'tool': 'Polyspace',
                'timestamp': datetime.now().isoformat(),
                'summary': summary,
                'violations': violations
            }, f, indent=2)
        logger.info(f"Saved violations to: {violations_file}")
        
        # Step 3: Create/Update Knowledge Base
        logger.info("STEP 3: Creating knowledge base...")
        kb_manager = KnowledgeDatabaseManager(args.module_name, kb_dir)
        
        # Convert violations to KB format
        for violation in violations:
            kb_manager.add_violation(
                violation_id=violation['violation_id'],
                violation_text=violation['violation_text'],
                severity=violation['severity'],
                category=violation['category'],
                file_path=violation['file'],
                line=violation['line']
            )
        
        kb_file = kb_manager.save_database()
        logger.info(f"Knowledge base saved: {kb_file}")
        
        # Step 4: Generate Excel report
        logger.info("STEP 4: Generating Excel report...")
        excel_report = reports_dir / f"{args.module_name}_violations_report.xlsx"
        generate_excel_report(violations, summary, excel_report, args.module_name)
        logger.info(f"Excel report saved: {excel_report}")
        
        # Step 5: Summary
        logger.info("="*80)
        logger.info("✅ Polyspace analysis completed successfully!")
        logger.info("="*80)
        logger.info("Output files:")
        logger.info(f"  📊 Excel Report: {excel_report}")
        logger.info(f"  💾 Knowledge Base: {kb_file}")
        logger.info(f"  📁 Violations JSON: {violations_file}")
        logger.info("="*80)
        
        if not args.no_fixes:
            logger.info("💡 Next step: Generate code fixes using Generate_Code_Fixes.bat")
        
        return 0
        
    except Exception as e:
        logger.error(f"[ERROR] Analysis failed: {e}", exc_info=True)
        return 1


def generate_excel_report(violations, summary, output_path, module_name):
    """Generate Excel report from Polyspace violations"""
    
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create DataFrame
        df_data = []
        for v in violations:
            df_data.append({
                'Violation ID': v['violation_id'],
                'Check ID': v['check_id'],
                'Color': v['color'],
                'Severity': v['severity'],
                'Category': v['category'],
                'Description': v['violation_text'],
                'File': v['file'],
                'Line': v['line'],
                'Comment': v.get('comment', ''),
                'Tool': 'Polyspace'
            })
        
        df = pd.DataFrame(df_data)
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # All violations sheet
            df.to_excel(writer, sheet_name='All Violations', index=False)
            
            # Red findings (proven errors)
            red_df = df[df['Color'] == 'red']
            if not red_df.empty:
                red_df.to_excel(writer, sheet_name='Red - Proven Errors', index=False)
            
            # Orange findings (potential errors)
            orange_df = df[df['Color'] == 'orange']
            if not orange_df.empty:
                orange_df.to_excel(writer, sheet_name='Orange - Potential Errors', index=False)
            
            # Summary sheet
            summary_data = [
                ['Metric', 'Count'],
                ['Total Findings', summary['total_findings']],
                ['Red (Proven Errors)', summary['red_findings']],
                ['Orange (Potential Errors)', summary['orange_findings']],
                ['Green (Proven Safe)', summary['green_findings']],
                ['Gray (Unreachable)', summary['gray_findings']],
                ['', ''],
                ['Category Breakdown', '']
            ]
            
            for category, count in summary['by_category'].items():
                summary_data.append([category, count])
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
        
        # Apply formatting
        wb = load_workbook(output_path)
        
        # Format all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Header row formatting
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header
            ws.freeze_panes = 'A2'
        
        wb.save(output_path)
        logger.info(f"Excel report formatted and saved")
        
    except ImportError as e:
        logger.warning(f"Could not generate Excel report: {e}")
        logger.warning("Install pandas and openpyxl: pip install pandas openpyxl")
    except Exception as e:
        logger.error(f"Failed to generate Excel report: {e}")


if __name__ == '__main__':
    sys.exit(main())
