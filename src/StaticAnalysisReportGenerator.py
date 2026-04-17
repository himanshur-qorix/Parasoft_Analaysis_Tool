"""
Static Analysis Report Generator
Generates color-coded HTML and Excel reports for static code analysis

Developer: Himanshu R
Organization: Qorix India Pvt Ltd
Version: 3.0.0
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict
import json

logger = logging.getLogger(__name__)


class StaticAnalysisReportGenerator:
    """Generate HTML and Excel reports for static code analysis"""
    
    def __init__(self, module_name: str, violations: List[Dict], summary: Dict):
        """
        Initialize report generator
        
        Args:
            module_name: Module name
            violations: List of violation dictionaries
            summary: Summary statistics
        """
        self.module_name = module_name
        self.violations = violations
        self.summary = summary
    
    def generate_html_report(self, output_path: Path) -> Path:
        """
        Generate HTML report with color-coded violations
        
        Args:
            output_path: Path for output HTML file
        
        Returns:
            Path to generated HTML file
        """
        try:
            logger.info(f"Generating HTML report: {output_path}")
            
            # Separate violations by color code
            red_violations = [v for v in self.violations if v.get('color_code') == 'RED']
            orange_violations = [v for v in self.violations if v.get('color_code') == 'ORANGE']
            grey_violations = [v for v in self.violations if v.get('color_code') == 'GREY']
            
            logger.info(f"Color breakdown: RED={len(red_violations)}, ORANGE={len(orange_violations)}, GREY={len(grey_violations)}")
            
            html_content = self._generate_html_content(red_violations, orange_violations, grey_violations)
            
            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"HTML report generated successfully: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to generate HTML report: {e}", exc_info=True)
            raise
    
    def _generate_html_content(self, red_violations, orange_violations, grey_violations) -> str:
        """Generate complete HTML content"""
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Static Code Analysis Report - {self.module_name}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            color: #333;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 10px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
        }}
        
        .header p {{
            font-size: 1.1em;
            opacity: 0.9;
        }}
        
        .logo {{
            font-size: 2em;
            margin-bottom: 10px;
            font-weight: bold;
            letter-spacing: 3px;
        }}
        
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            padding: 40px;
            background: #f8f9fa;
        }}
        
        .summary-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            text-align: center;
        }}
        
        .summary-card h3 {{
            color: #666;
            font-size: 0.9em;
            text-transform: uppercase;
            margin-bottom: 10px;
        }}
        
        .summary-card .value {{
            font-size: 2.5em;
            font-weight: bold;
            color: #667eea;
        }}
        
        .color-breakdown {{
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 20px;
            padding: 40px;
        }}
        
        .color-card {{
            background: white;
            border-left: 5px solid;
            padding: 20px;
            border-radius: 5px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        
        .color-card.red {{
            border-left-color: #dc3545;
        }}
        
        .color-card.orange {{
            border-left-color: #fd7e14;
        }}
        
        .color-card.grey {{
            border-left-color: #6c757d;
        }}
        
        .color-card h2 {{
            font-size: 1.5em;
            margin-bottom: 10px;
        }}
        
        .color-card .count {{
            font-size: 3em;
            font-weight: bold;
            margin: 10px 0;
        }}
        
        .color-card.red .count {{
            color: #dc3545;
        }}
        
        .color-card.orange .count {{
            color: #fd7e14;
        }}
        
        .color-card.grey .count {{
            color: #6c757d;
        }}
        
        .violations-section {{
            padding: 40px;
        }}
        
        .section-header {{
            font-size: 2em;
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 3px solid;
            display: flex;
            align-items: center;
            gap: 15px;
        }}
        
        .section-header.red {{
            color: #dc3545;
            border-bottom-color: #dc3545;
        }}
        
        .section-header.orange {{
            color: #fd7e14;
            border-bottom-color: #fd7e14;
        }}
        
        .section-header.grey {{
            color: #6c757d;
            border-bottom-color: #6c757d;
        }}
        
        .color-indicator {{
            width: 30px;
            height: 30px;
            border-radius: 50%;
            display: inline-block;
        }}
        
        .violation-card {{
            background: white;
            border: 1px solid #e0e0e0;
            border-left: 5px solid;
            border-radius: 5px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 5px rgba(0,0,0,0.05);
            transition: transform 0.2s, box-shadow 0.2s;
        }}
        
        .violation-card:hover {{
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
        }}
        
        .violation-card.red {{
            border-left-color: #dc3545;
        }}
        
        .violation-card.orange {{
            border-left-color: #fd7e14;
        }}
        
        .violation-card.grey {{
            border-left-color: #6c757d;
        }}
        
        .violation-header {{
            display: flex;
            justify-content: space-between;
            align-items: start;
            margin-bottom: 15px;
        }}
        
        .violation-id {{
            font-family: 'Courier New', monospace;
            font-size: 0.9em;
            color: #666;
            background: #f8f9fa;
            padding: 5px 10px;
            border-radius: 3px;
        }}
        
        .violation-severity {{
            display: inline-block;
            padding: 5px 15px;
            border-radius: 15px;
            font-size: 0.85em;
            font-weight: bold;
            color: white;
        }}
        
        .violation-severity.red {{
            background: #dc3545;
        }}
        
        .violation-severity.orange {{
            background: #fd7e14;
        }}
        
        .violation-severity.grey {{
            background: #6c757d;
        }}
        
        .violation-title {{
            font-size: 1.2em;
            font-weight: bold;
            margin-bottom: 10px;
            color: #333;
        }}
        
        .violation-meta {{
            display: flex;
            gap: 20px;
            margin-bottom: 15px;
            color: #666;
            font-size: 0.9em;
        }}
        
        .violation-meta span {{
            display: flex;
            align-items: center;
            gap: 5px;
        }}
        
        .code-snippet {{
            background: #1e1e1e;
            color: #d4d4d4;
            padding: 20px;
            border-radius: 5px;
            font-family: 'Courier New', Consolas, monospace;
            font-size: 0.9em;
            line-height: 1.6;
            overflow-x: auto;
            margin-top: 15px;
        }}
        
        .code-line {{
            white-space: pre;
            margin: 2px 0;
        }}
        
        .code-line.highlight {{
            background: #3a3d41;
            border-left: 3px solid #dc3545;
            padding-left: 10px;
            margin-left: -13px;
        }}
        
        .code-line .line-number {{
            color: #858585;
            margin-right: 15px;
            user-select: none;
        }}
        
        .no-violations {{
            text-align: center;
            padding: 40px;
            color: #28a745;
            font-size: 1.2em;
        }}
        
        .footer {{
            background: #f8f9fa;
            padding: 30px;
            text-align: center;
            color: #666;
            border-top: 1px solid #e0e0e0;
        }}
        
        .footer img {{
            height: 40px;
            margin-bottom: 10px;
        }}
        
        .tabs {{
            display: flex;
            gap: 10px;
            padding: 20px 40px 0;
            border-bottom: 2px solid #e0e0e0;
        }}
        
        .tab {{
            padding: 10px 20px;
            cursor: pointer;
            border: none;
            background: transparent;
            font-size: 1em;
            color: #666;
            border-bottom: 3px solid transparent;
            transition: all 0.3s;
        }}
        
        .tab:hover {{
            color: #667eea;
        }}
        
        .tab.active {{
            color: #667eea;
            border-bottom-color: #667eea;
            font-weight: bold;
        }}
        
        .tab-content {{
            display: none;
        }}
        
        .tab-content.active {{
            display: block;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="logo">QORIX</div>
            <h1>Static Code Analysis Report</h1>
            <p>Module: {self.module_name}</p>
            <p>Generated: {timestamp}</p>
        </div>
        
        <div class="summary">
            <div class="summary-card">
                <h3>Files Analyzed</h3>
                <div class="value">{self.summary['files_analyzed']}</div>
            </div>
            <div class="summary-card">
                <h3>Lines Analyzed</h3>
                <div class="value">{self.summary['lines_analyzed']:,}</div>
            </div>
            <div class="summary-card">
                <h3>Total Findings</h3>
                <div class="value">{self.summary['total_findings']}</div>
            </div>
        </div>
        
        <div class="color-breakdown">
            <div class="color-card red">
                <h2>RED CODE</h2>
                <div class="count">{len(red_violations)}</div>
                <p>Critical Issues - Proven Errors</p>
                <p style="margin-top: 10px; color: #666;">Null pointer dereference, division by zero</p>
            </div>
            <div class="color-card orange">
                <h2>ORANGE CODE</h2>
                <div class="count">{len(orange_violations)}</div>
                <p>High Priority - Likely Errors</p>
                <p style="margin-top: 10px; color: #666;">Uninitialized variables, buffer overflows</p>
            </div>
            <div class="color-card grey">
                <h2>GREY CODE</h2>
                <div class="count">{len(grey_violations)}</div>
                <p>Code Quality Issues</p>
                <p style="margin-top: 10px; color: #666;">MISRA/CERT violations, style issues</p>
            </div>
        </div>
        
        <div class="tabs">
            <button class="tab active" onclick="showTab('all')">All Violations ({self.summary['total_findings']})</button>
            <button class="tab" onclick="showTab('red')">RED ({len(red_violations)})</button>
            <button class="tab" onclick="showTab('orange')">ORANGE ({len(orange_violations)})</button>
            <button class="tab" onclick="showTab('grey')">GREY ({len(grey_violations)})</button>
        </div>
        
        <div id="all-tab" class="tab-content active violations-section">
            {self._generate_violations_html(red_violations, 'red', 'RED CODE - Critical Issues')}
            {self._generate_violations_html(orange_violations, 'orange', 'ORANGE CODE - High Priority')}
            {self._generate_violations_html(grey_violations, 'grey', 'GREY CODE - Code Quality')}
        </div>
        
        <div id="red-tab" class="tab-content violations-section">
            {self._generate_violations_html(red_violations, 'red', 'RED CODE - Critical Issues')}
        </div>
        
        <div id="orange-tab" class="tab-content violations-section">
            {self._generate_violations_html(orange_violations, 'orange', 'ORANGE CODE - High Priority')}
        </div>
        
        <div id="grey-tab" class="tab-content violations-section">
            {self._generate_violations_html(grey_violations, 'grey', 'GREY CODE - Code Quality')}
        </div>
        
        <div class="footer">
            <p><strong>Qorix India Pvt Ltd</strong></p>
            <p>Static Code Analyzer v3.0.0</p>
            <p style="margin-top: 10px; font-size: 0.9em;">
                Developer: Himanshu R | Organization: Qorix India Pvt Ltd
            </p>
        </div>
    </div>
    
    <script>
        function showTab(tabName) {{
            // Hide all tabs
            document.querySelectorAll('.tab-content').forEach(tab => {{
                tab.classList.remove('active');
            }});
            
            // Remove active from all tab buttons
            document.querySelectorAll('.tab').forEach(btn => {{
                btn.classList.remove('active');
            }});
            
            // Show selected tab
            document.getElementById(tabName + '-tab').classList.add('active');
            
            // Activate button
            event.target.classList.add('active');
        }}
    </script>
</body>
</html>
"""
        return html
    
    def _generate_violations_html(self, violations: List[Dict], color: str, title: str) -> str:
        """Generate HTML for a list of violations"""
        
        if not violations:
            return f'<div class="no-violations">[OK] No {color.upper()} violations found!</div>'
        
        color_indicator_style = {
            'red': '#dc3545',
            'orange': '#fd7e14',
            'grey': '#6c757d'
        }
        
        violations_html = f"""
        <div class="section-header {color}">
            <span class="color-indicator" style="background: {color_indicator_style[color]};"></span>
            {title} ({len(violations)})
        </div>
        """
        
        for v in violations:
            code_context_html = self._generate_code_context_html(v.get('code_context', []))
            
            violations_html += f"""
        <div class="violation-card {color}">
            <div class="violation-header">
                <div>
                    <div class="violation-id">{v['violation_id']}</div>
                    <div class="violation-title">{v['check_id']}</div>
                </div>
                <span class="violation-severity {color}">{v['severity']}</span>
            </div>
            
            <div class="violation-meta">
                <span>File: {v['file']}</span>
                <span>Line {v['line']}</span>
                <span>Category: {v['category']}</span>
            </div>
            
            <p style="margin-bottom: 15px; color: #555;">{v['violation_text']}</p>
            
            <div class="code-snippet">
                {code_context_html}
            </div>
        </div>
        """
        
        return violations_html
    
    def _generate_code_context_html(self, code_context: List[Dict]) -> str:
        """Generate HTML for code context with highlighting"""
        
        if not code_context:
            return '<div class="code-line">No code context available</div>'
        
        html_lines = []
        for ctx in code_context:
            line_class = 'code-line highlight' if ctx['is_violation'] else 'code-line'
            line_num = ctx['line_num']
            content = ctx['content'].replace('<', '&lt;').replace('>', '&gt;')
            marker = '>>> ' if ctx['is_violation'] else '    '
            
            html_lines.append(
                f'<div class="{line_class}"><span class="line-number">{marker}{line_num:4d}:</span>{content}</div>'
            )
        
        return '\n'.join(html_lines)
    
    def generate_excel_report(self, output_path: Path) -> Path:
        """
        Generate Excel report with enhanced code snippets
        
        Args:
            output_path: Path for output Excel file
        
        Returns:
            Path to generated Excel file
        """
        try:
            logger.info(f"Generating Excel report: {output_path}")
            
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Separate by color code
            red_violations = [v for v in self.violations if v.get('color_code') == 'RED']
            orange_violations = [v for v in self.violations if v.get('color_code') == 'ORANGE']
            grey_violations = [v for v in self.violations if v.get('color_code') == 'GREY']
            
            logger.info(f"Excel color breakdown: RED={len(red_violations)}, ORANGE={len(orange_violations)}, GREY={len(grey_violations)}")
            
            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Create DataFrame with multi-line code snippets
            df_data = []
            for v in self.violations:
                df_data.append({
                    'Violation ID': v['violation_id'],
                    'Color Code': v['color_code'],
                    'Check ID': v['check_id'],
                    'Severity': v['severity'],
                    'Category': v['category'],
                    'Description': v['violation_text'],
                    'File': v['file'],
                    'Line': v['line'],
                    'Code Snippet (Single Line)': v.get('code_snippet', ''),
                    'Code Context (Multi-line)': v.get('code_snippet_multiline', ''),
                    'Tool': 'StaticAnalyzer'
                })
            
            df = pd.DataFrame(df_data)
            
            # Create Excel file with multiple sheets
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # All violations
                df.to_excel(writer, sheet_name='All Violations', index=False)
                
                # RED violations
                if red_violations:
                    red_df = pd.DataFrame([{
                        'Violation ID': v['violation_id'],
                        'Check ID': v['check_id'],
                        'Description': v['violation_text'],
                        'File': v['file'],
                        'Line': v['line'],
                        'Code Context': v.get('code_snippet_multiline', '')
                    } for v in red_violations])
                    red_df.to_excel(writer, sheet_name='RED - Critical', index=False)
                
                # ORANGE violations
                if orange_violations:
                    orange_df = pd.DataFrame([{
                        'Violation ID': v['violation_id'],
                        'Check ID': v['check_id'],
                        'Description': v['violation_text'],
                        'File': v['file'],
                        'Line': v['line'],
                        'Code Context': v.get('code_snippet_multiline', '')
                    } for v in orange_violations])
                    orange_df.to_excel(writer, sheet_name='ORANGE - High Priority', index=False)
                
                # GREY violations
                if grey_violations:
                    grey_df = pd.DataFrame([{
                        'Violation ID': v['violation_id'],
                        'Check ID': v['check_id'],
                        'Description': v['violation_text'],
                        'File': v['file'],
                        'Line': v['line'],
                        'Code Context': v.get('code_snippet_multiline', '')
                    } for v in grey_violations])
                    grey_df.to_excel(writer, sheet_name='GREY - Code Quality', index=False)
                
                # By category sheets
                for category in df['Category'].unique():
                    cat_df = df[df['Category'] == category]
                    sheet_name = f"{category[:25]}"  # Excel limit with space for prefix
                    cat_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Summary
                summary_data = [
                    ['Static Code Analysis Summary', ''],
                    ['Module', self.module_name],
                    ['Generated', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                    ['', ''],
                    ['Analysis Statistics', ''],
                    ['Files Analyzed', self.summary['files_analyzed']],
                    ['Lines Analyzed', self.summary['lines_analyzed']],
                    ['Total Findings', self.summary['total_findings']],
                    ['', ''],
                    ['Color Code Breakdown', ''],
                    ['RED (Critical)', len(red_violations)],
                    ['ORANGE (High)', len(orange_violations)],
                    ['GREY (Quality)', len(grey_violations)],
                    ['', ''],
                    ['Severity Breakdown', ''],
                    ['Critical', self.summary['critical']],
                    ['High', self.summary['high']],
                    ['Medium', self.summary['medium']],
                    ['Low', self.summary['low']],
                    ['', ''],
                    ['Category Breakdown', '']
                ]
                
                for category, count in self.summary['by_category'].items():
                    summary_data.append([category, count])
                
                summary_data.extend([['', ''], ['Top Violation Types', '']])
                sorted_checks = sorted(self.summary['by_check'].items(), key=lambda x: x[1], reverse=True)
                for check, count in sorted_checks[:15]:
                    summary_data.append([check, count])
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
            
            # Apply formatting
            wb = load_workbook(output_path)
            
            # Color fills for severity
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            orange_fill = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')
            grey_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            
            # Format sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Header row formatting
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Apply color coding to rows based on color code column
                if sheet_name == 'All Violations' and ws.max_row > 1:
                    for row in range(2, ws.max_row + 1):
                        color_code_cell = ws.cell(row=row, column=2)  # Column B (Color Code)
                        
                        if color_code_cell.value == 'RED':
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = red_fill
                        elif color_code_cell.value == 'ORANGE':
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = orange_fill
                        elif color_code_cell.value == 'GREY':
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = grey_fill
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            cell_length = len(str(cell.value)) if cell.value else 0
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 80)  # Cap at 80
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Make code context column monospace
                if 'Code Context' in [cell.value for cell in ws[1]]:
                    code_col_idx = [cell.value for cell in ws[1]].index('Code Context') + 1
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=code_col_idx).font = Font(name='Courier New', size=9)
                        ws.cell(row=row, column=code_col_idx).alignment = Alignment(wrap_text=True, vertical='top')
            
            wb.save(output_path)
            logger.info(f"Excel report generated: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}", exc_info=True)
            raise
