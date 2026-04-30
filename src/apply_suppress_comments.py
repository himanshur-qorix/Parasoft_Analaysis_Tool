"""
Apply Suppress Comments to Source Code
Interactive tool to insert Parasoft suppress comments into source files
Version: 1.0.0
Developer: Himanshu R
"""

import sys
import re
import json
from pathlib import Path
from datetime import datetime
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure UTF-8 encoding for Windows console
if sys.platform == 'win32':
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleOutputCP(65001)  # UTF-8
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except:
        pass

class SuppressCommentApplicator:
    """Applies Parasoft suppress comments to source code with user confirmation"""
    
    def __init__(self, suppress_file_path, target_repo_path):
        """
        Initialize the applicator
        
        Args:
            suppress_file_path: Path to the suppress comments file
            target_repo_path: Path to the target repository with source code
        """
        self.suppress_file = Path(suppress_file_path)
        self.target_repo = Path(target_repo_path)
        self.applied_count = 0
        self.skipped_count = 0
        self.failed_count = 0
        self.already_justified_count = 0  # NEW: Track already justified lines
        self.skipped_in_comment_count = 0  # NEW: Track skipped due to comment blocks
        self.applied_violations = []  # Track successfully applied violations (file, line)
        self.skipped_violations = []  # Track user-rejected violations (file, line)
        
        # Extract module name from suppress file (e.g., "Mka" from "Mka_suppress_comments_20260410_112659.txt")
        self.module_name = self.suppress_file.stem.split('_suppress_comments_')[0]
        
        if not self.suppress_file.exists():
            raise FileNotFoundError(f"Suppress comments file not found: {suppress_file_path}")
        
        if not self.target_repo.exists():
            raise FileNotFoundError(f"Target repository not found: {target_repo_path}")
        
        # Create backup folder in project root (not in target repo to avoid search conflicts)
        # Get the project root (parent of scripts folder)
        project_root = Path(__file__).parent.parent
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.backup_folder = project_root / "backups" / f"parasoft_backups_{timestamp}"
        self.backup_folder.mkdir(parents=True, exist_ok=True)
        
        print(f"\n{'='*80}")
        print("  PARASOFT SUPPRESS COMMENTS APPLICATOR")
        print(f"{'='*80}\n")
        print(f"Module:            {self.module_name}")
        print(f"Suppress file:     {self.suppress_file}")
        print(f"Target repository: {self.target_repo}")
        print(f"Backup folder:     {self.backup_folder}\n")
    
    def parse_suppress_file(self):
        """
        Parse the suppress comments file (NEW INLINE FORMAT)
        
        Returns:
            Dictionary with:
                - 'files': Dict of {filename: {
                    'reference_section': str,
                    'suppressions': [{line: int, comment: str}, ...]
                }}
        """
        suppressions_data = {}
        
        with open(self.suppress_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Check if this is new inline format or old format
        if "NEW INLINE FORMAT:" in content:
            # NEW FORMAT: Parse by file sections
            file_sections = re.split(r'={80}\nFILE: (.+?)\n={80}\n', content)
            
            # Skip header (first element)
            for i in range(1, len(file_sections), 2):
                if i + 1 >= len(file_sections):
                    break
                    
                file_name = file_sections[i].strip()
                file_content = file_sections[i + 1]
                
                # Extract reference section
                ref_match = re.search(
                    r'--- VIOLATIONS REFERENCE SECTION.*?\n(.*?)(?=--- INLINE SUPPRESS COMMENTS|$)',
                    file_content,
                    re.DOTALL
                )
                reference_section = ref_match.group(1).strip() if ref_match else ""
                
                # Extract inline suppressions
                suppressions = []
                inline_pattern = r'Line (\d+):\n\s+// parasoft-suppress (.+)'
                for match in re.finditer(inline_pattern, file_content):
                    line_num = int(match.group(1))
                    suppress_comment = f"// parasoft-suppress {match.group(2).strip()}"
                    suppressions.append({
                        'line': line_num,
                        'comment': suppress_comment
                    })
                
                suppressions_data[file_name] = {
                    'reference_section': reference_section,
                    'suppressions': suppressions
                }
        else:
            # OLD FORMAT: Legacy support
            pattern = r'File: (.+?), Line: (\d+)\s*-+\s*(/\* parasoft-begin-suppress .+? \*/)\s*.+?\s*(/\* parasoft-end-suppress .+? \*/)'
            matches = re.finditer(pattern, content, re.DOTALL)
            
            for match in matches:
                file_name = match.group(1).strip()
                line_num = int(match.group(2))
                begin_comment = match.group(3).strip()
                end_comment = match.group(4).strip()
                
                if file_name not in suppressions_data:
                    suppressions_data[file_name] = {
                        'reference_section': '',
                        'suppressions': []
                    }
                
                suppressions_data[file_name]['suppressions'].append({
                    'file': file_name,
                    'line': line_num,
                    'begin_comment': begin_comment,
                    'end_comment': end_comment
                })
        
        return suppressions_data
    
    def find_source_file(self, file_name):
        """
        Find a source file in the target repository
        
        Args:
            file_name: Name of the file to find
        
        Returns:
            Path to the file or None if not found
        """
        # Search recursively in target repository, excluding backup folders
        all_matches = list(self.target_repo.rglob(file_name))
        
        # Filter out any files in backup folders
        matches = [
            m for m in all_matches 
            if 'parasoft_backup' not in str(m).lower() and 'backup' not in m.parts
        ]
        
        if len(matches) == 0:
            return None
        elif len(matches) == 1:
            return matches[0]
        else:
            # Multiple matches - let user choose
            print(f"\n[WARNING] Multiple files named '{file_name}' found:")
            for i, path in enumerate(matches, 1):
                print(f"  {i}. {path}")
            
            while True:
                choice = input(f"Select file (1-{len(matches)}) or 's' to skip: ").strip().lower()
                if choice == 's':
                    return None
                try:
                    idx = int(choice) - 1
                    if 0 <= idx < len(matches):
                        return matches[idx]
                except ValueError:
                    pass
                print("Invalid choice. Try again.")
    
    def create_backup(self, file_path):
        """
        Create a backup of the file before modification
        
        Args:
            file_path: Path to the file to backup
        
        Returns:
            Path to the backup file
        """
        # Get relative path from target repo
        try:
            rel_path = file_path.relative_to(self.target_repo)
        except ValueError:
            # If file is not under target_repo, use just the filename
            rel_path = Path(file_path.name)
        
        # Create backup path maintaining directory structure
        backup_path = self.backup_folder / rel_path
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Copy file to backup location
        shutil.copy2(file_path, backup_path)
        return backup_path
    
    def is_inside_comment_block(self, lines, line_num):
        """
        Check if a line is inside a comment block
        
        Args:
            lines: List of file lines
            line_num: Line number to check (1-indexed)
        
        Returns:
            True if inside comment block, False otherwise
        """
        # Track if we're inside a multi-line comment
        in_comment = False
        
        for i in range(line_num):
            line = lines[i].strip()
            
            # Check for comment start
            if '/*' in line:
                in_comment = True
            
            # Check for comment end
            if '*/' in line:
                in_comment = False
        
        return in_comment
    
    def is_already_suppressed(self, lines, line_num):
        """
        Check if a line already has suppress comments around it
        
        Args:
            lines: List of file lines
            line_num: Line number to check (1-indexed)
        
        Returns:
            True if suppress comments already exist, False otherwise
        """
        if line_num < 1 or line_num > len(lines):
            return False
        
        # Check a few lines before and after for parasoft suppress comments
        search_range_before = max(0, line_num - 5)
        search_range_after = min(len(lines), line_num + 5)
        
        has_begin = False
        has_end = False
        
        # Look for begin-suppress before the line
        for i in range(search_range_before, line_num):
            if 'parasoft-begin-suppress' in lines[i]:
                has_begin = True
                break
        
        # Look for end-suppress after the line
        for i in range(line_num - 1, search_range_after):
            if 'parasoft-end-suppress' in lines[i]:
                has_end = True
                break
        
        return has_begin and has_end
    
    def show_preview(self, file_path, line_num, begin_comment, end_comment):
        """
        Show a preview of the changes
        
        Args:
            file_path: Path to the source file
            line_num: Line number where suppression applies
            begin_comment: Begin suppress comment
            end_comment: End suppress comment
        
        Returns:
            Tuple of (current line content, is_in_comment, is_already_suppressed)
        """
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        
        # Check if line is inside a comment block
        is_in_comment = self.is_inside_comment_block(lines, line_num)
        
        # Check if already has suppress comments
        is_already_suppressed = self.is_already_suppressed(lines, line_num)
        
        # Show context around the line
        start_idx = max(0, line_num - 3)
        end_idx = min(len(lines), line_num + 3)
        
        print(f"\n{'-'*80}")
        print(f"File: {file_path.name}")
        print(f"Line: {line_num}")
        
        if is_already_suppressed:
            print("[INFO] This line already has suppress comments!")
        elif is_in_comment:
            print("[WARNING] This line appears to be inside a comment block!")
        
        print(f"{'-'*80}")
        print("\nCurrent code:")
        for i in range(start_idx, end_idx):
            marker = ">>>" if i == line_num - 1 else "   "
            print(f"{marker} {i+1:4d}: {lines[i].rstrip()}")
        
        if is_already_suppressed:
            print("\n[SKIP] Suppress comments already exist for this line")
        elif not is_in_comment:
            print("\nProposed changes:")
            print(f"     ADD: {begin_comment}")
            if line_num - 1 < len(lines):
                print(f"     {line_num:4d}: {lines[line_num - 1].rstrip()}")
            print(f"     ADD: {end_comment}")
        else:
            print("\n[SKIP RECOMMENDED] Cannot insert suppress comments inside comment blocks")
        
        current_line = lines[line_num - 1] if line_num - 1 < len(lines) else ""
        return current_line, is_in_comment, is_already_suppressed
    
    def apply_suppression(self, file_path, line_num, begin_comment, end_comment):
        """
        Apply suppression comments to a file
        
        Args:
            file_path: Path to the source file
            line_num: Line number where suppression applies
            begin_comment: Begin suppress comment
            end_comment: End suppress comment
        
        Note: This method is deprecated in favor of apply_inline_suppression for the new format
        """
        # This method is kept for backward compatibility but not used by new format
        pass
    
    def apply_inline_suppression(self, file_path, line_num, suppress_comment):
        """
        Apply inline suppression comment to the end of a line
        
        Args:
            file_path: Path to the file
            line_num: Line number to apply suppression (1-indexed)
            suppress_comment: Inline comment to add
        
        Returns:
            True if successful, False otherwise
        """
        try:
            # Read file
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            
            if line_num < 1 or line_num > len(lines):
                print(f"[ERROR] Line number {line_num} is out of range")
                return False
            
            # Create backup
            backup_path = self.create_backup(file_path)
            print(f"[INFO] Backup created: {backup_path.name}")
            
            # Get the line (0-indexed)
            line_idx = line_num - 1
            line = lines[line_idx].rstrip('\n\r')
            
            # Check if already has suppress comment
            if 'parasoft-suppress' in line:
                print(f"[INFO] Line already has suppress comment - skipping")
                return False
            
            # Add inline comment at the end of the line
            modified_line = f"{line} {suppress_comment}\n"
            lines[line_idx] = modified_line
            
            # Write modified file
            with open(file_path, 'w', encoding='utf-8', newline='') as f:
                f.writelines(lines)
            
            return True
            
        except Exception as e:
            print(f"[ERROR] Failed to apply inline suppression: {str(e)}")
            return False
    
    def add_violations_reference_section(self, file_path, reference_section):
        """
        Add or update violations reference section at top of file
        
        Args:
            file_path: Path to the file
            reference_section: Reference section content to add
        
        Returns:
            True if successful, False otherwise
        """
        try:
            # Read file
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            
            # Find existing violations section
            section_start = -1
            section_end = -1
            
            for i, line in enumerate(lines):
                if 'Parasoft violations Section' in line:
                    # Find the start of the comment block
                    for j in range(i, -1, -1):
                        if '/*'  in lines[j] and'*****' in lines[j]:
                            section_start = j
                            break
                    # Find the end (the closing */ of the section)
                    for j in range(i, len(lines)):
                        if '*/' in lines[j] and section_start == -1:
                            continue
                        if '*/' in lines[j]:
                            section_end = j
                            break
                    break
            
            #  Parse and insert new reference section
            new_section_lines = reference_section.split('\n')
            new_section_lines = [line + '\n' if not line.endswith('\n') else line 
                               for line in new_section_lines]
            
            if section_start != -1 and section_end != -1:
                # Replace existing section
                print("[INFO] Updating existing violations reference section")
                lines[section_start:section_end+1] = new_section_lines
            else:
                # Add new section after existing header comments
                print("[INFO] Adding new violations reference section")
                insert_pos = 0
                
                # Find position after copyright/file header
                in_header = False
                for i, line in enumerate(lines):
                    if '/*' in line and '***' in line:
                        in_header = True
                    if in_header and '*/' in line:
                        insert_pos = i + 1
                        break
                
                # Insert new section
                new_section_lines.append('\n')  # Add blank line after
                for line in reversed(new_section_lines):
                    lines.insert(insert_pos, line)
            
            # Write modified file
            with open(file_path, 'w', encoding='utf-8', newline='') as f:
                f.writelines(lines)
            
            return True
            
        except Exception as e:
            print(f"[ERROR] Failed to add reference section: {str(e)}")
            return False
    
    def run(self, auto_apply=False):
        """Run the interactive application process (supports both old and new formats)
        
        Args:
            auto_apply: If True, apply all suppressions without prompting
        """
        
        # Parse suppress file
        print("Parsing suppress comments file...")
        suppressions_data = self.parse_suppress_file()
        
        if not suppressions_data:
            print("\n[WARNING] No suppressions found in file")
            return
        
        total_files = len(suppressions_data)
        total_suppressions = sum(len(data['suppressions']) for data in suppressions_data.values())
        print(f"Found {total_suppressions} suppression(s) across {total_files} file(s)\n")
        
        if auto_apply:
            print("[INFO] 🚀 AUTO-APPLY MODE - Applying all suppressions without prompting\n")
        
        # Process each file
        for file_name, file_data in suppressions_data.items():
            print(f"\n{'='*80}")
            print(f"Processing File: {file_name}")
            print(f"{'='*80}")
            
            # Find source file
            source_file = self.find_source_file(file_name)
            
            if not source_file:
                print(f"[ERROR] Could not find file: {file_name}")
                self.failed_count += len(file_data['suppressions'])
                continue
            
            # Add violations reference section if present
            if file_data['reference_section']:
                print("\n[INFO] Adding violations reference section to file...")
                if self.add_violations_reference_section(source_file, file_data['reference_section']):
                    print("[SUCCESS] Reference section added/updated")
                else:
                    print("[WARNING] Failed to add reference section")
            
            # Process suppressions for this file
            suppressions = file_data['suppressions']
            
            for i, supp in enumerate(suppressions, 1):
                print(f"\n{'-'*80}")
                print(f"Suppression {i} of {len(suppressions)} in {file_name}")
                print(f"{'-'*80}")
                
                # Check if this is inline or old format
                if 'comment' in supp:
                    # NEW INLINE FORMAT
                    line_num = supp['line']
                    suppress_comment = supp['comment']
                    
                    # Show preview
                    try:
                        with open(source_file, 'r', encoding='utf-8', errors='ignore') as f:
                            lines = f.readlines()
                        
                        if line_num < 1 or line_num > len(lines):
                            print(f"[ERROR] Line {line_num} out of range")
                            self.failed_count += 1
                            continue
                        
                        # Show context
                        start_idx = max(0, line_num - 3)
                        end_idx = min(len(lines), line_num + 3)
                        
                        print(f"\nLine {line_num}:")
                        for j in range(start_idx, end_idx):
                            prefix = ">>> " if j == line_num - 1 else "    "
                            print(f"{prefix}{j+1:4}: {lines[j].rstrip()}")
                        
                        print(f"\nWill add: {suppress_comment}")
                        
                        # Check if already has suppression
                        if 'parasoft-suppress' in lines[line_num - 1]:
                            print("\n[AUTO-SKIP] Line already has suppress comment")
                            self.already_justified_count += 1
                            continue
                        
                        # Ask for confirmation (unless auto_apply mode)
                        if auto_apply:
                            choice = 'y'  # Auto-apply
                        else:
                            choice = input(f"\nApply inline suppression? (y/n/a=all/q=quit): ").strip().lower()
                        
                        if choice == 'y':
                            if self.apply_inline_suppression(source_file, line_num, suppress_comment):
                                print("[SUCCESS] Inline suppression applied")
                                self.applied_count += 1
                                self.applied_violations.append({'file': file_name, 'line': line_num})
                            else:
                                self.failed_count += 1
                        
                        elif choice == 'a':
                            # Apply this and all remaining in this file
                            if self.apply_inline_suppression(source_file, line_num, suppress_comment):
                                print("[SUCCESS] Inline suppression applied")
                                self.applied_count += 1
                                self.applied_violations.append({'file': file_name, 'line': line_num})
                            else:
                                self.failed_count += 1
                            
                            # Apply remaining
                            for remaining_supp in suppressions[i:]:
                                rem_line = remaining_supp['line']
                                rem_comment = remaining_supp['comment']
                                
                                # Re-read file for fresh line check
                                with open(source_file, 'r', encoding='utf-8', errors='ignore') as f:
                                    fresh_lines = f.readlines()
                                
                                if rem_line >= len(fresh_lines) or 'parasoft-suppress' in fresh_lines[rem_line - 1]:
                                    print(f"[AUTO-SKIP] Line {rem_line} - already has suppress comment")
                                    self.already_justified_count += 1
                                elif self.apply_inline_suppression(source_file, rem_line, rem_comment):
                                    print(f"[SUCCESS] Applied suppression to line {rem_line}")
                                    self.applied_count += 1
                                    self.applied_violations.append({'file': file_name, 'line': rem_line})
                                else:
                                    self.failed_count += 1
                            break  # Exit loop after applying all
                        
                        elif choice == 'q':
                            print("\n[INFO] User quit")
                            self.show_summary()
                            return
                        
                        else:
                            print("[INFO] Skipped")
                            self.skipped_count += 1
                            self.skipped_violations.append({'file': file_name, 'line': line_num})
                    
                    except Exception as e:
                        print(f"[ERROR] {str(e)}")
                        self.failed_count += 1
                
                else:
                    # OLD FORMAT (Legacy support)
                    line_num = supp['line']
                    begin_comment = supp['begin_comment']
                    end_comment = supp['end_comment']
                    
                    # Use old method
                    current_line, is_in_comment, is_already_suppressed = self.show_preview(
                        source_file, line_num, begin_comment, end_comment
                    )
                    
                    if is_already_suppressed:
                        print("\n[AUTO-SKIP] Already has suppress comments")
                        self.already_justified_count += 1
                        continue
                    
                    if is_in_comment:
                        print("\n[AUTO-SKIP] Inside comment block")
                        self.skipped_in_comment_count += 1
                        continue
                    
                    # Ask for confirmation (unless auto_apply mode)
                    if auto_apply:
                        choice = 'y'  # Auto-apply
                    else:
                        choice = input(f"\nApply suppression? (y/n/q): ").strip().lower()
                    
                    if choice == 'y':
                        if self.apply_suppression(source_file, line_num, begin_comment, end_comment):
                            print("[SUCCESS] Suppression applied")
                            self.applied_count += 1
                            self.applied_violations.append({'file': file_name, 'line': line_num})
                        else:
                            self.failed_count += 1
                    elif choice == 'q':
                        print("\n[INFO] User quit")
                        self.show_summary()
                        return
                    else:
                        print("[INFO] Skipped")
                        self.skipped_count += 1
                        self.skipped_violations.append({'file': file_name, 'line': line_num})
        
        # Update reports and show summary
        self.update_module_violations_report()
        self.update_violations_report()
        self.show_summary()
    
    def update_module_violations_report(self):
        """Update the module-specific violations report with 'Justification updated' status"""
        if not self.applied_violations and not self.skipped_violations:
            print("\n[INFO] No violations were applied or skipped - skipping module report update")
            return
        
        # Get the project root (parent of scripts folder)
        project_root = Path(__file__).parent.parent
        
        # Construct module report filename (check both regular and _UPDATED versions)
        module_report_name = f"{self.module_name}_violations_report.xlsx"
        module_report_name_updated = f"{self.module_name}_violations_report_UPDATED.xlsx"
        
        # Look for module-specific report (prioritize _UPDATED version)
        possible_paths = [
            project_root / "reports" / module_report_name_updated,
            Path.cwd() / "reports" / module_report_name_updated,
            Path.cwd() / module_report_name_updated,
            project_root / module_report_name_updated,
            project_root / "reports" / module_report_name,
            Path.cwd() / "reports" / module_report_name,
            Path.cwd() / module_report_name,
            project_root / module_report_name
        ]
        
        report_file = None
        print(f"\n[INFO] Searching for {module_report_name} or {module_report_name_updated}...")
        for path in possible_paths:
            if path.exists():
                report_file = path
                print(f"[SUCCESS] Found at: {path}")
                break
        
        if not report_file:
            print(f"[WARNING] Module report not found")
            print(f"[INFO] Searched for: {module_report_name} or {module_report_name_updated}")
            print(f"[INFO] Skipping module violations report update")
            return
        
        try:
            print(f"[INFO] Updating module violations report...")
            
            # Load the workbook
            workbook = load_workbook(report_file)
            
            # Find the "Detailed Violations" sheet (or first sheet with violation data)
            ws = None
            for sheet_name in workbook.sheetnames:
                if "Detailed" in sheet_name or "Violations" in sheet_name:
                    test_ws = workbook[sheet_name]
                    # Check if it has File and Line number columns
                    headers_row = [cell.value for cell in test_ws[1]]
                    if "File" in headers_row and ("Line number" in headers_row or "Line" in headers_row):
                        ws = test_ws
                        print(f"[INFO] Found violations data in sheet: {sheet_name}")
                        break
            
            if ws is None:
                print("[WARNING] Could not find violations data sheet with 'File' and 'Line number' columns")
                return
            
            # Find header row and columns
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value] = col_idx
            
            print(f"[DEBUG] Found columns: {list(headers.keys())}")
            
            # Check if "Tool Update" column exists, if not add it after "Status"
            if "Tool Update" not in headers:
                # Find Status column
                status_col = headers.get("Status")
                if status_col:
                    tool_update_col = status_col + 1
                    # Insert new column
                    ws.insert_cols(tool_update_col)
                    ws.cell(row=1, column=tool_update_col, value="Tool Update")
                    headers["Tool Update"] = tool_update_col
                    print(f"[INFO] Added 'Tool Update' column after 'Status' at column {tool_update_col}")
                else:
                    # Add at the end
                    tool_update_col = max(headers.values()) + 1
                    ws.cell(row=1, column=tool_update_col, value="Tool Update")
                    headers["Tool Update"] = tool_update_col
                    print(f"[INFO] Added 'Tool Update' column at end (column {tool_update_col})")
            else:
                tool_update_col = headers["Tool Update"]
                print(f"[INFO] Using existing 'Tool Update' column at column {tool_update_col}")
            
            # Get File, Line number, and Justification column indices
            file_col = headers.get("File")
            line_col = headers.get("Line number") or headers.get("Line")
            justification_col = headers.get("Justification")
            
            if not file_col or not line_col:
                print("[WARNING] Could not find 'File' or 'Line number' columns")
                print(f"[DEBUG] Available columns: {list(headers.keys())}")
                return
            
            if not justification_col:
                print("[WARNING] Could not find 'Justification' column - will only update 'Tool Update'")
            
            print(f"[DEBUG] File column: {file_col}, Line column: {line_col}, Tool Update column: {tool_update_col}, Justification column: {justification_col}")
            
            # Update rows for applied violations
            updated_count = 0
            rejected_count = 0
            print(f"[DEBUG] Checking {ws.max_row - 1} rows for matches...")
            print(f"[DEBUG] Looking for {len(self.applied_violations)} applied violations:")
            for av in self.applied_violations[:5]:  # Show first 5
                print(f"  - {av['file']}:{av['line']}")
            if len(self.applied_violations) > 5:
                print(f"  ... and {len(self.applied_violations) - 5} more")
            
            if self.skipped_violations:
                print(f"[DEBUG] Looking for {len(self.skipped_violations)} skipped violations to mark as 'No'")
            
            for row_idx in range(2, ws.max_row + 1):
                file_val = ws.cell(row=row_idx, column=file_col).value
                line_val = ws.cell(row=row_idx, column=line_col).value
                
                if file_val and line_val:
                    # Check if this violation was applied (accepted)
                    for applied in self.applied_violations:
                        # Match by filename (not full path) and line number
                        if applied['file'] in str(file_val) and int(line_val) == int(applied['line']):
                            # Update Tool Update column
                            ws.cell(row=row_idx, column=tool_update_col, value="Justification updated")
                            
                            # Update Justification column to "Yes" (accepted)
                            if justification_col:
                                ws.cell(row=row_idx, column=justification_col, value="Yes")
                            
                            updated_count += 1
                            if updated_count <= 3:  # Show first 3 matches
                                print(f"[DEBUG] Matched (Applied): {file_val}:{line_val} (Row {row_idx})")
                            break
                    
                    # Check if this violation was skipped (rejected)
                    for skipped in self.skipped_violations:
                        if skipped['file'] in str(file_val) and int(line_val) == int(skipped['line']):
                            # Update Justification column to "No" (rejected)
                            if justification_col:
                                ws.cell(row=row_idx, column=justification_col, value="No")
                            
                            rejected_count += 1
                            if rejected_count <= 3:  # Show first 3 matches
                                print(f"[DEBUG] Matched (Skipped): {file_val}:{line_val} (Row {row_idx})")
                            break
            
            # Save workbook
            workbook.save(report_file)
            print(f"[SUCCESS] Updated {updated_count} violations in {report_file.name}")
            if justification_col:
                print(f"[INFO] Set 'Justification' to 'Yes' for {updated_count} applied violations")
                if rejected_count > 0:
                    print(f"[INFO] Set 'Justification' to 'No' for {rejected_count} rejected violations")
            
        except PermissionError:
            print(f"\n[ERROR] Cannot update {report_file.name if report_file else module_report_name} - file is open")
            print("[TIP] Please close the Excel file and run the tool again")
        except Exception as e:
            print(f"\n[WARNING] Could not update {report_file.name if report_file else module_report_name}: {e}")
            import traceback
            traceback.print_exc()
    
    def update_violations_report(self):
        """Update the violations_report Excel file with suppression statistics"""
        # Get the project root (parent of scripts folder)
        project_root = Path(__file__).parent.parent
        
        # Look for violations_report.xlsx in common locations
        possible_paths = [
            Path.cwd() / "violations_report.xlsx",
            Path.cwd() / "data" / "violations_report.xlsx",
            project_root / "violations_report.xlsx",
            project_root / "data" / "violations_report.xlsx",
            self.target_repo / "violations_report.xlsx",
            self.target_repo.parent / "violations_report.xlsx",
            self.suppress_file.parent.parent / "violations_report.xlsx"
        ]
        
        report_file = None
        print("\n[DEBUG] Searching for violations_report.xlsx in:")
        for path in possible_paths:
            print(f"  - {path}")
            if path.exists():
                report_file = path
                print(f"  ✓ Found at: {path}")
                break
        
        if not report_file:
            print("\n[INFO] violations_report.xlsx not found in any expected location")
            print("[INFO] Skipping Excel update")
            print("[TIP] Make sure violations_report.xlsx exists in project root or data folder")
            return
        
        try:
            print(f"\n[INFO] Updating violations report: {report_file}")
            
            # Load the workbook
            workbook = load_workbook(report_file)
            
            # Create or update Suppression Summary sheet
            if "Suppression Summary" in workbook.sheetnames:
                # Append to existing sheet instead of replacing
                ws = workbook["Suppression Summary"]
                # Find next row
                next_row = ws.max_row + 2  # Add blank row before new entry
            else:
                # Create new summary sheet
                ws = workbook.create_sheet("Suppression Summary", 0)
                next_row = 1
            
            # Prepare summary data
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            summary_df = pd.DataFrame([
                {"Metric": "Module", "Value": self.module_name},
                {"Metric": "Suppress File", "Value": self.suppress_file.name},
                {"Metric": "Timestamp", "Value": timestamp},
                {"Metric": "Applied", "Value": self.applied_count},
                {"Metric": "Skipped (User)", "Value": self.skipped_count},
                {"Metric": "Already Justified", "Value": self.already_justified_count},
                {"Metric": "Skipped (In Comment)", "Value": self.skipped_in_comment_count},
                {"Metric": "Failed", "Value": self.failed_count},
                {"Metric": "Total", "Value": self.applied_count + self.skipped_count + self.already_justified_count + self.skipped_in_comment_count + self.failed_count},
                {"Metric": "Backup Folder", "Value": str(self.backup_folder) if self.applied_count > 0 else "N/A"}
            ])
            
            # Write to sheet starting from next_row
            for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=(next_row == 1)), next_row):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            workbook.save(report_file)
            print(f"[SUCCESS] Updated violations_report with suppression summary")
            
        except PermissionError:
            print(f"\n[ERROR] Cannot update violations_report.xlsx - file is open")
            print("[TIP] Please close the Excel file and run the tool again")
        except Exception as e:
            print(f"\n[WARNING] Could not update violations_report.xlsx: {e}")
            import traceback
            traceback.print_exc()
    
    def show_summary(self):
        """Show summary of operations and save to JSON"""
        total = self.applied_count + self.skipped_count + self.already_justified_count + self.skipped_in_comment_count + self.failed_count
        
        print(f"\n{'='*80}")
        print("  SUPPRESSION SUMMARY")
        print(f"{'='*80}")
        print(f"Module:               {self.module_name}")
        print(f"File:                 {self.suppress_file.name}")
        print(f"Applied:              {self.applied_count}")
        print(f"Skipped (user):       {self.skipped_count}")
        print(f"Already Justified:    {self.already_justified_count}")
        print(f"Skipped (in comment): {self.skipped_in_comment_count}")
        print(f"Failed:               {self.failed_count}")
        print(f"{'='*80}")
        print(f"Total:                {total}")
        print(f"{'='*80}")
        
        if self.applied_count > 0:
            print(f"\nBackup files saved to: {self.backup_folder}")
            print("Review backups if you need to restore original files")
            print(f"\nModule report status: Check console output above for update details")
            print(f"  - {len(self.applied_violations)} violations applied and tracked")
        
        # Save summary to JSON for reporting
        summary_data = {
            "module": self.module_name,
            "suppress_file": self.suppress_file.name,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "applied": self.applied_count,
            "skipped_user": self.skipped_count,
            "already_justified": self.already_justified_count,
            "skipped_in_comment": self.skipped_in_comment_count,
            "failed": self.failed_count,
            "total": total,
            "backup_folder": str(self.backup_folder) if self.applied_count > 0 else None,
            "applied_violations": self.applied_violations,
            "skipped_violations": self.skipped_violations
        }
        
        # Save to justifications folder
        summary_file = Path(self.suppress_file).parent / f"suppression_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        try:
            with open(summary_file, 'w') as f:
                json.dump(summary_data, f, indent=2)
            print(f"\nSummary saved to: {summary_file}")
        except Exception as e:
            print(f"\n[WARNING] Could not save summary file: {e}")
        
        print()


def main():
    """Main entry point"""
    
    if len(sys.argv) < 3:
        print("\n" + "="*80)
        print("  PARASOFT SUPPRESSION COMMENT APPLICATOR")
        print("="*80)
        print("\nUsage: python apply_suppress_comments.py <suppress_file> <target_repo> [--auto-apply]")
        print("\nArguments:")
        print("  suppress_file : Path to suppression comments file (justifications/*.txt)")
        print("  target_repo   : Path to target source code repository")
        print("  --auto-apply  : (Optional) Apply all suppressions without prompting")
        print("\nExamples:")
        print("  Interactive mode (prompts y/n/a for each suppression):")
        print("    python apply_suppress_comments.py justifications\\Mka_suppress_comments_20260410_143022.txt D:\\MyProject\\src")
        print("\n  Auto-apply mode (applies all without prompting):")
        print("    python apply_suppress_comments.py justifications\\Mka_suppress_comments_20260410_143022.txt D:\\MyProject\\src --auto-apply")
        print("\n" + "="*80 + "\n")
        sys.exit(1)
    
    suppress_file = sys.argv[1]
    target_repo = sys.argv[2]
    auto_apply = "--auto-apply" in sys.argv or "-a" in sys.argv
    
    if auto_apply:
        print("\n[INFO] 🚀 AUTO-APPLY MODE: All suppressions will be applied without prompting")
        print("[INFO] Press Ctrl+C within 2 seconds to cancel\n")
        import time
        time.sleep(2)  # Give user time to cancel
    
    try:
        applicator = SuppressCommentApplicator(suppress_file, target_repo)
        applicator.run(auto_apply=auto_apply)
    except FileNotFoundError as e:
        print(f"\n[ERROR] {str(e)}")
        sys.exit(1)
    except KeyboardInterrupt:
        print("\n\n[INFO] Operation cancelled by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n[ERROR] Unexpected error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
