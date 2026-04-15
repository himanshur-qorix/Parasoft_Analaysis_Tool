"""
Code Fix Generator
Generates code fixes and justifications for Parasoft violations
Enhanced with Parasoft Rules Database integration

Developer: Himanshu R
Organization: Qorix India Pvt Ltd
Version: 2.0.0
"""

import logging
import json
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
from KnowledgeDatabaseManager import KnowledgeDatabaseManager
from OllamaIntegration import OllamaIntegration
from ParasoftRulesParser import ParasoftRulesParser

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed - Excel update features disabled")


class CodeFixGenerator:
    """Generates fixes and justifications for Parasoft violations"""
    
    def __init__(self, module_name: str, kb_manager: KnowledgeDatabaseManager, 
                 fixes_dir: Path, config: Optional[Dict] = None):
        """
        Initialize the Code Fix Generator
        
        Args:
            module_name: Name of the module
            kb_manager: Knowledge database manager instance
            fixes_dir: Directory to store generated fixes
            config: Optional configuration dictionary
        """
        self.module_name = module_name
        self.kb_manager = kb_manager
        self.kb_manager.load_knowledge_base(module_name)
        self.fixes_dir = Path(fixes_dir)
        self.fixes_dir.mkdir(exist_ok=True)
        
        # Create module-specific fixes directory
        # If fixes_dir already contains module name, use it directly
        if module_name in str(fixes_dir):
            self.module_fixes_dir = self.fixes_dir
        else:
            self.module_fixes_dir = self.fixes_dir / module_name
            self.module_fixes_dir.mkdir(exist_ok=True)
        
        # Initialize Ollama integration
        if config is None:
            # Load config from file
            config_path = Path(__file__).parent.parent / 'config' / 'config.json'
            if config_path.exists():
                with open(config_path, 'r') as f:
                    config = json.load(f)
            else:
                config = {}
        
        # Initialize Parasoft Rules Database (PRIMARY SOURCE)
        rules_dir = Path(__file__).parent.parent / 'data' / 'Parasoft_Enabled_Rules_List' / 'gendoc'
        if rules_dir.exists():
            try:
                self.rules_parser = ParasoftRulesParser(rules_dir)
                rules_count = self.rules_parser.load_all_rules()
                logger.info(f"[OK] Parasoft Rules Database loaded: {rules_count} rules available")
                self.use_rules_db = True
            except Exception as e:
                logger.warning(f"Failed to load Parasoft Rules Database: {e}")
                self.rules_parser = None
                self.use_rules_db = False
        else:
            logger.warning(f"Parasoft Rules directory not found: {rules_dir}")
            self.rules_parser = None
            self.use_rules_db = False
        
        ai_config = config.get('ai_integration', {})
        self.ollama = OllamaIntegration(ai_config)
        
        # Log AI status
        status = self.ollama.get_status()
        ai_mode = status.get('ai_mode', 'hybrid')
        if status['enabled']:
            logger.info(f"[OK] AI Mode: {ai_mode} | Model: {status['model']} at {status['base_url']}")
        else:
            if ai_mode == 'rules_only':
                logger.info(f"[INFO] AI Mode: {ai_mode} - Using rule-based fixes only")
            else:
                logger.info(f"[INFO] AI disabled - Using rule-based fixes")
        
        # Store source code path if provided
        self.source_code_path = config.get('source_code_path')
        if self.source_code_path:
            self.source_code_path = Path(self.source_code_path)
            logger.info(f"[INFO] Source code path configured: {self.source_code_path}")
        
        # Store config for later use
        self.config = config
        
        logger.info(f"Code Fix Generator initialized for module: {module_name}")
    
    def generate_all_fixes(self, violation_ids: Optional[List[str]] = None) -> Dict:
        """
        Generate fixes for violations
        
        Args:
            violation_ids: List of specific violation IDs (None = all unfixed)
        
        Returns:
            Results dictionary
        """
        if violation_ids:
            violations_to_fix = [
                self.kb_manager.get_violation(vid)
                for vid in violation_ids
                if self.kb_manager.get_violation(vid)
            ]
        else:
            violations_to_fix = [
                v for v in self.kb_manager.get_all_violations()
                if not v.get('fix_applied')
            ]
        
        logger.info(f"Generating fixes for {len(violations_to_fix)} violations")
        
        fixes_generated = 0
        fixes_failed = 0
        
        all_fixes = []
        justification_recommendations = []  # Track AI justification suggestions
        
        for violation in violations_to_fix:
            try:
                fix_data = self._generate_fix_for_violation(violation)
                if fix_data:
                    all_fixes.append(fix_data)
                    fixes_generated += 1
                    
                    # Collect justification recommendations
                    if fix_data.get('recommended_action') == 'justify':
                        justification_recommendations.append(fix_data)
                else:
                    fixes_failed += 1
            except Exception as e:
                logger.error(f"Error generating fix for {violation['violation_id']}: {str(e)}")
                fixes_failed += 1
        
        # Save all fixes to text and HTML files
        if all_fixes:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            fixes_file = self.module_fixes_dir / f"{self.module_name}_fixes_{timestamp}.txt"
            html_file = self.module_fixes_dir / f"{self.module_name}_fixes_{timestamp}.html"
            
            self._save_fixes_file(all_fixes, fixes_file)
            self.save_fixes_html(all_fixes, html_file)
            
            logger.info(f"Fixes saved to: {fixes_file}")
            logger.info(f"HTML report saved to: {html_file}")
        
        # Add justification recommendations to Excel report
        if justification_recommendations:
            logger.info(f"Found {len(justification_recommendations)} AI justification recommendations")
            excel_updated = self._add_justification_sheet_to_excel(justification_recommendations)
            if excel_updated:
                logger.info(f"✅ Added 'AI Justification Suggestions' sheet to violations report")
            else:
                logger.warning(f"⚠️  Could not update Excel report - file may not exist yet")
        
        return {
            'fixes_generated': fixes_generated,
            'fixes_failed': fixes_failed,
            'total_violations': len(violations_to_fix),
            'fixes_file': str(fixes_file) if all_fixes else None,
            'justification_recommendations': len(justification_recommendations),
            'html_file': str(html_file) if all_fixes else None
        }
    
    def _generate_fix_for_violation(self, violation: Dict) -> Optional[Dict]:
        """
        Generate fix for a specific violation
        
        Args:
            violation: Violation dictionary
        
        Returns:
            Fix data dictionary
        """
        violation_id = violation['violation_id']
        violation_text = violation['violation_text']
        severity = violation['severity']
        category = violation['category']
        
        # STEP 1: Check cross-module handling to see if this should be justified instead
        cross_module_info = self._check_cross_module_justifications(violation_id)
        
        # STEP 2: If other modules justified this, ask AI if it should be justified here too
        justification_suggestion = None
        if cross_module_info and self.ollama.enabled:
            justification_suggestion = self.ollama.suggest_justification(violation, cross_module_info)
            
            if justification_suggestion and justification_suggestion.get('should_justify'):
                confidence = justification_suggestion.get('confidence', 'MEDIUM')
                reason = justification_suggestion.get('reason', 'Common deviation')
                
                logger.info(f"⚖️  AI suggests JUSTIFICATION for {violation_id} ({confidence} confidence)")
                logger.info(f"   Reason: {reason}")
                logger.info(f"   Cross-module context: {len(cross_module_info)} modules have justified this")
                
                # Return special fix data suggesting justification
                return {
                    'violation_id': violation_id,
                    'violation_text': violation_text,
                    'severity': severity,
                    'category': category,
                    'files_affected': violation['files_affected'],
                    'fix_suggestion': {
                        'type': 'justification_recommended',
                        'description': f"AI Analysis: This violation appears to be a common deviation. {len(cross_module_info)} other module(s) have justified this same violation. Consider adding to common deviations Excel instead of generating code fix.",
                        'example': f"Suggestion: Add {violation_id} to common deviations list\n\nReason: {justification_suggestion.get('suggested_rationale', reason)}\n\nModules with justifications:\n" + "\n".join([f"  - {info}" for info in cross_module_info[:5]]),
                        'priority': 'HIGH' if confidence == 'HIGH' else 'MEDIUM',
                        'ai_generated': True,
                        'justification_analysis': justification_suggestion
                    },
                    'code_context': None,
                    'timestamp': datetime.now().isoformat(),
                    'recommended_action': 'justify'
                }
        
        # STEP 3: If no justification recommended, proceed with normal fix generation
        # Extract code context if source code path is provided
        code_context = None
        if self.source_code_path:
            code_context = self._extract_code_context(violation)
        
        # Generate fix based on violation type
        fix_suggestion = self._get_fix_suggestion(violation_id, violation_text, category, code_context)
        
        if not fix_suggestion:
            logger.warning(f"No fix suggestion available for {violation_id}")
            return None
        
        fix_data = {
            'violation_id': violation_id,
            'violation_text': violation_text,
            'severity': severity,
            'category': category,
            'files_affected': violation['files_affected'],
            'fix_suggestion': fix_suggestion,
            'code_context': code_context,
            'timestamp': datetime.now().isoformat()
        }
        
        # Update knowledge base
        self.kb_manager.update_fix_status(violation_id, {
            'fix_generated': True,
            'fix_type': fix_suggestion['type'],
            'timestamp': datetime.now().isoformat()
        })
        
        return fix_data
    
    def _check_cross_module_justifications(self, violation_id: str) -> list:
        """
        Check if other modules have justified the same violation
        
        Args:
            violation_id: Violation ID to check
        
        Returns:
            List of informational strings about modules that justified this violation
        """
        justified_in_modules = []
        
        try:
            # Get all knowledge bases
            kb_dir = self.kb_manager.kb_dir
            kb_files = list(kb_dir.glob('*_KnowledgeDatabase.json'))
            
            # Extract base violation ID (remove trailing instance numbers)
            import re
            base_violation_id = re.sub(r'-\d+$', '', violation_id)
            
            for kb_file in kb_files:
                # Skip current module
                if kb_file.stem == f"{self.module_name}_KnowledgeDatabase":
                    continue
                
                try:
                    import json
                    with open(kb_file, 'r', encoding='utf-8') as f:
                        kb_data = json.load(f)
                    
                    module_name = kb_file.stem.replace('_KnowledgeDatabase', '')
                    violations = kb_data.get('violations', {})
                    
                    # Check for exact or base match
                    for vid, vdata in violations.items():
                        if vid == violation_id or vid.startswith(base_violation_id):
                            # Check if it was justified/suppressed
                            if vdata.get('justification_added') or vdata.get('suppressed'):
                                justified_count = len(vdata.get('files_affected', []))
                                justified_in_modules.append(
                                    f"Module '{module_name}': {justified_count} instance(s) justified/suppressed"
                                )
                                break  # Found in this module, move to next KB
                
                except Exception as e:
                    logger.debug(f"Error reading KB file {kb_file}: {e}")
                    continue
        
        except Exception as e:
            logger.error(f"Error checking cross-module justifications: {e}")
        
        return justified_in_modules
    
    def _extract_code_context(self, violation: Dict) -> Optional[Dict]:
        """
        Extract code context from source file for better fix suggestions
        
        Args:
            violation: Violation dictionary with file and line info
        
        Returns:
            Dictionary with code context or None if not available
        """
        if not self.source_code_path:
            return None
        
        try:
            files_affected = violation.get('files_affected', [])
            if not files_affected:
                return None
            
            # Parse first file entry (format: "file.c:line")
            file_entry = files_affected[0]
            if ':' in file_entry:
                file_name, line_str = file_entry.rsplit(':', 1)
                try:
                    line_number = int(line_str)
                except ValueError:
                    return None
            else:
                return None
            
            # Search for the file in source code path
            source_file = self._find_source_file(file_name)
            if not source_file:
                return None
            
            # Read code context (5 lines before, target line, 5 lines after)
            context_lines = 5
            with open(source_file, 'r', encoding='utf-8', errors='ignore') as f:
                all_lines = f.readlines()
            
            total_lines = len(all_lines)
            start_line = max(0, line_number - context_lines - 1)
            end_line = min(total_lines, line_number + context_lines)
            
            context_code = ''.join(all_lines[start_line:end_line])
            target_line_code = all_lines[line_number - 1] if 0 < line_number <= total_lines else ''
            
            return {
                'file': file_name,
                'line': line_number,
                'target_code': target_line_code.strip(),
                'context': context_code,
                'start_line': start_line + 1,
                'end_line': end_line
            }
        
        except Exception as e:
            logger.debug(f"Error extracting code context: {e}")
            return None
    
    def _find_source_file(self, filename: str) -> Optional[Path]:
        """
        Find source file in the source code directory
        
        Args:
            filename: File name to search for
        
        Returns:
            Path to file or None if not found
        """
        if not self.source_code_path or not self.source_code_path.exists():
            return None
        
        # Search recursively for the file
        for path in self.source_code_path.rglob(filename):
            if path.is_file():
                return path
        
        return None
    
    def _get_fix_suggestion(self, violation_id: str, violation_text: str, category: str, code_context: Optional[Dict] = None) -> Optional[Dict]:
        """
        Get fix suggestion based on AI mode configuration
        
        Strategy depends on AI mode:
        - AI Only: AIgeneration -> Rule-based fallback (skip Parasoft DB)
        - Hybrid: Parasoft DB -> AI -> Rule-based fallback (balanced approach)
        - Rules Only: Parasoft DB -> Rule-based patterns (skip AI)
        
        Args:
            violation_id: Violation ID
            violation_text: Violation description
            category: Violation category
            code_context: Optional code context from source file
        
        Returns:
            Fix suggestion dictionary
        """
        # Get AI mode to determine strategy
        ai_mode = self.ollama.ai_mode
        
        # AI ONLY MODE: Use AI with Parasoft DB as reference context
        if ai_mode == 'ai_only':
            logger.debug(f"[AI-ONLY] Attempting AI generation for {violation_id}")
            
            # Get Parasoft DB information as reference (not as direct answer)
            parasoft_reference = None
            if self.use_rules_db and self.rules_parser:
                parasoft_reference = self._get_parasoft_reference(violation_id, violation_text, category)
                if parasoft_reference:
                    logger.debug(f"[AI-ONLY] Found Parasoft reference for context")
            
            # Try AI generation with enhanced context
            if self.ollama.enabled:
                violation_dict = {
                    'violation_id': violation_id,
                    'violation_text': violation_text,
                    'category': category,
                    'severity': 'MEDIUM',
                    'code_context': code_context,
                    'parasoft_reference': parasoft_reference
                }
                
                ai_fix = self.ollama.generate_fix_suggestion(violation_dict)
                if ai_fix:
                    logger.info(f"[AI-ONLY] Using AI-generated fix for {violation_id}")
                    return ai_fix
            
            # Fallback to rule-based only if AI fails
            logger.info(f"[AI-ONLY] AI unavailable, falling back to rules for {violation_id}")
            return self._get_rule_based_fix(violation_id, violation_text, category)
        
        # RULES ONLY MODE: Use Parasoft DB and patterns, skip AI
        elif ai_mode == 'rules_only':
            logger.debug(f"[RULES-ONLY] Using Parasoft DB + patterns for {violation_id}")
            
            # Try Parasoft Rules Database first
            if self.use_rules_db and self.rules_parser:
                parasoft_fix = self._get_parasoft_official_fix(violation_id, violation_text, category)
                if parasoft_fix:
                    logger.info(f"[PARASOFT-DB] Using official Parasoft fix for {violation_id}")
                    return parasoft_fix
            
            # Fallback to rule-based patterns
            logger.info(f"[RULES-ONLY] Using pattern-based fix for {violation_id}")
            return self._get_rule_based_fix(violation_id, violation_text, category)
        
        # HYBRID MODE (default): Try all sources in priority order
        else:
            logger.debug(f"[HYBRID] Trying all sources for {violation_id}")
            
            # PRIORITY 1: Try Parasoft Rules Database for official fix
            if self.use_rules_db and self.rules_parser:
                parasoft_fix = self._get_parasoft_official_fix(violation_id, violation_text, category)
                if parasoft_fix:
                    logger.info(f"[PARASOFT-DB] Using official Parasoft fix for {violation_id}")
                    return parasoft_fix
            
            # PRIORITY 2: Try AI generation (if appropriate)
            if self.ollama.should_use_ai(category, violation_text):
                violation_dict = {
                    'violation_id': violation_id,
                    'violation_text': violation_text,
                    'category': category,
                    'severity': 'MEDIUM',
                    'code_context': code_context
                }
                
                ai_fix = self.ollama.generate_fix_suggestion(violation_dict)
                if ai_fix:
                    logger.info(f"[AI] Using AI-generated fix for {violation_id}")
                    return ai_fix
            
            # PRIORITY 3: Fallback to rule-based patterns
            logger.info(f"[HYBRID] Using pattern-based fix for {violation_id}")
            return self._get_rule_based_fix(violation_id, violation_text, category)
    
    def _get_rule_based_fix(self, violation_id: str, violation_text: str, category: str) -> Optional[Dict]:
        """
        Get rule-based fix using pattern matching
        
        Args:
            violation_id: Violation ID
            violation_text: Violation description
            category: Violation category
        
        Returns:
            Fix suggestion dictionary
        """
        text_upper = violation_text.upper()
        
        # Common MISRA fixes
        if 'MISRA' in category:
            return self._get_misra_fix(violation_id, text_upper, violation_text)
        
        # Common CERT fixes
        elif 'CERT' in category:
            return self._get_cert_fix(violation_id, text_upper, violation_text)
        
        # Generic fixes
        else:
            return self._get_generic_fix(text_upper, violation_text)
    
    def _get_parasoft_official_fix(self, violation_id: str, violation_text: str, category: str) -> Optional[Dict]:
        """
        Extract official fix from Parasoft Rules Database
        
        Args:
            violation_id: Violation ID (e.g., "CERT_C-STR31-a")
            violation_text: Violation description
            category: Violation category
        
        Returns:
            Fix suggestion dictionary with official Parasoft repair example, or None
        """
        try:
            # Extract base rule ID from violation ID
            # Example: "CERT_C-STR31-a-2" -> "CERT_C-STR31-a"
            # Example: "MISRAC2012-RULE_8_7-a" -> "MISRAC2012-RULE_8_7-a"
            
            # Try exact match first
            rule = self.rules_parser.get_rule(violation_id)
            
            # If not found, try extracting base rule (remove trailing numbers/variant)
            if not rule:
                import re
                # Try removing trailing -\d+ suffix
                base_id = re.sub(r'-\d+$', '', violation_id)
                rule = self.rules_parser.get_rule(base_id)
            
            if not rule:
                return None
            
            # Build fix suggestion from official Parasoft documentation
            fix_parts = []
            
            # Add description
            if rule.description:
                fix_parts.append(f"Description: {rule.description[:200]}...")
            
            # Add security relevance if available
            if rule.security_relevance and rule.security_relevance != "N/A":
                fix_parts.append(f"\nSecurity Impact: {rule.security_relevance[:150]}...")
            
            # Add repair example (MOST IMPORTANT)
            repair_code = ""
            if rule.example_repair:
                repair_code = rule.example_repair
                # Limit repair code length for display
                if len(repair_code) > 500:
                    repair_code = repair_code[:500] + "\n... (truncated)"
            
            # Add violation example for context
            violation_code = ""
            if rule.example_violation:
                violation_code = rule.example_violation
                if len(violation_code) > 300:
                    violation_code = violation_code[:300] + "\n... (truncated)"
            
            # Build formatted fix suggestion
            fix_description = "\n".join(fix_parts)
            
            example_text = ""
            if violation_code:
                example_text += f"\n--- VIOLATION EXAMPLE ---\n{violation_code}\n"
            if repair_code:
                example_text += f"\n--- OFFICIAL REPAIR ---\n{repair_code}\n"
            
            # Add CWE mappings if available
            cwe_info = ""
            if rule.cwe_mappings:
                cwe_info = f"CWE Mappings: {', '.join(rule.cwe_mappings[:5])}"
            
            return {
                'type': 'parasoft_official',
                'standard': rule.standard,
                'priority': 'HIGH',  # Official Parasoft fixes are high priority
                'description': fix_description,
                'example': example_text,
                'cwe_mappings': rule.cwe_mappings,
                'repair_code': repair_code,
                'violation_code': violation_code,
                'references': rule.references[:3] if rule.references else [],  # Limit to 3
                'source': f'Parasoft Official Documentation: {rule.title}'
            }
            
        except Exception as e:
            logger.debug(f"Could not extract Parasoft fix for {violation_id}: {e}")
            return None
    
    def _get_parasoft_reference(self, violation_id: str, violation_text: str, category: str) -> Optional[Dict]:
        """
        Get Parasoft DB information as reference context (not as direct fix)
        This is used to provide AI with official documentation for better suggestions
        
        Args:
            violation_id: Violation ID
            violation_text: Violation description
            category: Violation category
        
        Returns:
            Dictionary with Parasoft reference info or None
        """
        try:
            # Try exact match first
            rule = self.rules_parser.get_rule(violation_id)
            
            # If not found, try extracting base rule
            if not rule:
                import re
                base_id = re.sub(r'-\d+$', '', violation_id)
                rule = self.rules_parser.get_rule(base_id)
            
            if not rule:
                return None
            
            # Extract key information as reference context
            reference_info = {
                'rule_title': rule.title,
                'description': rule.description[:500] if rule.description else None,
                'security_relevance': rule.security_relevance[:300] if rule.security_relevance else None,
                'repair_strategy': rule.repair[:400] if rule.repair else None,
                'example_violation': rule.example_violation[:300] if rule.example_violation else None,
                'example_repair': rule.example_repair[:300] if rule.example_repair else None,
                'cwe_mappings': rule.cwe_mappings[:3] if rule.cwe_mappings else []
            }
            
            return reference_info
            
        except Exception as e:
            logger.debug(f"Could not extract Parasoft reference for {violation_id}: {e}")
            return None
    
    def _get_misra_fix(self, violation_id: str, text_upper: str, violation_text: str) -> Dict:
        """Get MISRA-specific fix suggestions"""
        
        # Rule 10.1 - Type conversions
        if '10.1' in violation_id or 'IMPLICIT CONVERSION' in text_upper:
            return {
                'type': 'explicit_cast',
                'description': 'Add explicit type cast to avoid implicit conversion',
                'example': '''
// Before:
uint8_t a = 5;
uint16_t b = a + 10;  // Implicit conversion

// After:
uint8_t a = 5;
uint16_t b = (uint16_t)a + 10U;  // Explicit cast
''',
                'priority': 'HIGH'
            }
        
        # Rule 11.x - Pointer conversions
        elif '11.' in violation_id or 'POINTER' in text_upper:
            return {
                'type': 'pointer_cast',
                'description': 'Use proper pointer casting or avoid incompatible pointer types',
                'example': '''
// Before:
int* ptr = (int*)some_void_ptr;

// After:
int* ptr = (int*)(uintptr_t)some_void_ptr;  // Use intermediate cast
''',
                'priority': 'HIGH'
            }
        
        # Rule 14.x - Control flow
        elif '14.' in violation_id or 'FOR LOOP' in text_upper or 'WHILE' in text_upper:
            return {
                'type': 'control_flow',
                'description': 'Ensure proper loop structure and counter usage',
                'example': '''
// Ensure loop counter is not modified within the loop body
// Use appropriate loop bounds and increment
''',
                'priority': 'MEDIUM'
            }
        
        # Rule 21.x - Standard library
        elif '21.' in violation_id or 'STANDARD LIBRARY' in text_upper:
            return {
                'type': 'stdlib_replacement',
                'description': 'Avoid or replace restricted standard library function',
                'example': '''
// Replace unsafe library functions with safe alternatives
// Example: Use strncpy instead of strcpy
''',
                'priority': 'HIGH'
            }
        
        # Default MISRA fix
        else:
            return {
                'type': 'misra_compliance',
                'description': f'Review MISRA rule and apply appropriate fix: {violation_text}',
                'example': 'Refer to MISRA C guidelines for specific rule details',
                'priority': 'MEDIUM'
            }
    
    def _get_cert_fix(self, violation_id: str, text_upper: str, violation_text: str) -> Dict:
        """Get CERT-specific fix suggestions"""
        
        # Buffer overflow issues
        if 'BUFFER' in text_upper or 'OVERFLOW' in text_upper:
            return {
                'type': 'buffer_safety',
                'description': 'Ensure buffer boundaries are checked before access',
                'example': '''
// Before:
char buffer[10];
strcpy(buffer, user_input);  // Unsafe

// After:
char buffer[10];
strncpy(buffer, user_input, sizeof(buffer) - 1);
buffer[sizeof(buffer) - 1] = '\\0';  // Ensure null termination
''',
                'priority': 'CRITICAL'
            }
        
        # Null pointer issues
        elif 'NULL' in text_upper or 'DEREFERENCE' in text_upper:
            return {
                'type': 'null_check',
                'description': 'Add null pointer check before dereference',
                'example': '''
// Before:
result = ptr->value;

// After:
if (ptr != NULL) {
    result = ptr->value;
} else {
    // Handle null pointer case
    result = default_value;
}
''',
                'priority': 'HIGH'
            }
        
        # Memory management
        elif 'MEMORY' in text_upper or 'LEAK' in text_upper or 'FREE' in text_upper:
            return {
                'type': 'memory_management',
                'description': 'Ensure proper memory allocation and deallocation',
                'example': '''
// Ensure every malloc has a corresponding free
// Check allocation success before use
void* ptr = malloc(size);
if (ptr != NULL) {
    // Use memory
    free(ptr);
    ptr = NULL;  // Prevent double-free
}
''',
                'priority': 'HIGH'
            }
        
        # Default CERT fix
        else:
            return {
                'type': 'cert_compliance',
                'description': f'Review CERT rule and apply appropriate fix: {violation_text}',
                'example': 'Refer to CERT Secure Coding Standards',
                'priority': 'MEDIUM'
            }
    
    def _get_generic_fix(self, text_upper: str, violation_text: str) -> Dict:
        """Get generic fix suggestions"""
        
        return {
            'type': 'generic',
            'description': violation_text,
            'example': 'Review the violation and apply appropriate fix based on context',
            'priority': 'MEDIUM'
        }
    
    def save_fixes_html(self, fixes: List[Dict], output_file: Path):
        """
        Save fixes to an HTML file with Qorix branding and syntax highlighting
        
        Args:
            fixes: List of fix dictionaries
            output_file: Path to output HTML file
        """
        from datetime import datetime
        
        html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Code Fix Suggestions - {self.module_name} - Qorix</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            color: #333;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}
        
        .header-logo {{
            width: 120px;
            height: 120px;
            margin: 0 auto 20px;
        }}
        
        .company {{
            font-size: 1.2em;
            opacity: 0.9;
            margin-top: 10px;
        }}
        
        h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
        }}
        
        .stats {{
            display: flex;
            justify-content: space-around;
            background: #f8f9fa;
            padding: 25px;
            margin: 0;
            border-bottom: 3px solid #667eea;
        }}
        
        .stat {{
            text-align: center;
        }}
        
        .stat-value {{
            font-size: 2.5em;
            font-weight: bold;
            color: #667eea;
        }}
        
        .stat-label {{
            color: #666;
            margin-top: 5px;
            font-size: 0.9em;
        }}
        
        .filters {{
            padding: 20px 40px;
            background: #f8f9fa;
            border-bottom: 1px solid #dee2e6;
            display: flex;
            gap: 15px;
            flex-wrap: wrap;
            align-items: center;
        }}
        
        .filter-group {{
            display: flex;
            gap: 10px;
            align-items: center;
        }}
        
        .filter-group label {{
            font-weight: 600;
            color: #495057;
        }}
        
        .filter-group input,
        .filter-group select {{
            padding: 8px 12px;
            border: 1px solid #ced4da;
            border-radius: 5px;
            font-size: 14px;
        }}
        
        .fixes-container {{
            padding: 40px;
        }}
        
        .fix-card {{
            background: white;
            border: 2px solid #e9ecef;
            border-radius: 10px;
            margin-bottom: 30px;
            overflow: hidden;
            transition: all 0.3s ease;
        }}
        
        .fix-card:hover {{
            border-color: #667eea;
            box-shadow: 0 5px 20px rgba(102, 126, 234, 0.2);
        }}
        
        .fix-header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px 25px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            cursor: pointer;
        }}
        
        .fix-title {{
            font-size: 1.3em;
            font-weight: 600;
        }}
        
        .fix-badges {{
            display: flex;
            gap: 10px;
        }}
        
        .badge {{
            padding: 5px 12px;
            border-radius: 20px;
            font-size: 0.85em;
            font-weight: 600;
            background: rgba(255,255,255,0.2);
        }}
        
        .badge-critical {{ background: #dc3545; }}
        .badge-high {{ background: #fd7e14; }}
        .badge-medium {{ background: #ffc107; color: #333; }}
        .badge-low {{ background: #28a745; }}
        
        .badge-parasoft {{ background: #17a2b8; }}
        .badge-ai {{ background: #6f42c1; }}
        .badge-rule {{ background: #6c757d; }}
        .badge-info {{ background: #20c997; }}
        
        /* Justification recommendation card styling */
        .justification-rec-card {{
            border-left: 5px solid #20c997 !important;
            background: linear-gradient(135deg, #e8f9f5 0%, #ffffff 100%) !important;
        }}
        
        .justification-rec-card .fix-header {{
            background: linear-gradient(135deg, #20c997 0%, #17a589 100%) !important;
        }}
        
        .fix-content {{
            padding: 25px;
        }}
        
        .section {{
            margin-bottom: 25px;
        }}
        
        .section-title {{
            font-size: 1.1em;
            font-weight: 600;
            color: #667eea;
            margin-bottom: 10px;
            padding-bottom: 5px;
            border-bottom: 2px solid #e9ecef;
        }}
        
        .description {{
            line-height: 1.6;
            color: #495057;
            background: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            border-left: 4px solid #667eea;
        }}
        
        .code-block {{
            background: #282c34;
            color: #abb2bf;
            padding: 20px;
            border-radius: 8px;
            overflow-x: auto;
            font-family: 'Consolas', 'Monaco', monospace;
            font-size: 14px;
            line-height: 1.5;
            margin: 10px 0;
        }}
        
        .code-header {{
            background: #21252b;
            color: #61afef;
            padding: 10px 15px;
            border-radius: 8px 8px 0 0;
            font-weight: 600;
            border-left: 4px solid #61afef;
        }}
        
        .violation-code {{
            border-left: 4px solid #e06c75;
        }}
        
        .repair-code {{
            border-left: 4px solid #98c379;
        }}
        
        .files-list {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            max-height: 200px;
            overflow-y: auto;
        }}
        
        .file-item {{
            padding: 8px 12px;
            margin: 5px 0;
            background: white;
            border-left: 3px solid #667eea;
            border-radius: 3px;
            font-family: monospace;
            font-size: 0.9em;
        }}
        
        .cwe-tags {{
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
            margin-top: 10px;
        }}
        
        .cwe-tag {{
            background: #e7f3ff;
            color: #0066cc;
            padding: 5px 12px;
            border-radius: 15px;
            font-size: 0.85em;
            font-weight: 600;
            border: 1px solid #99ccff;
        }}
        
        .footer {{
            background: #f8f9fa;
            padding: 30px;
            text-align: center;
            border-top: 3px solid #667eea;
            color: #666;
        }}
        
        .footer-logo {{
            width: 60px;
            height: 60px;
            margin: 0 auto 15px;
        }}
        
        @media print {{
            body {{
                background: white;
                padding: 0;
            }}
            .fix-card {{
                page-break-inside: avoid;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-logo">
                <svg viewBox="0 0 200 200" xmlns="http://www.w3.org/2000/svg">
                    <circle cx="100" cy="100" r="90" fill="white" opacity="0.2"/>
                    <text x="100" y="120" font-size="80" font-weight="bold" text-anchor="middle" fill="white">Q</text>
                </svg>
            </div>
            <h1>Code Fix Suggestions</h1>
            <div class="company">Qorix India Pvt Ltd</div>
            <p style="margin-top: 15px; opacity: 0.9;">Module: {self.module_name}</p>
            <p style="opacity: 0.8; font-size: 0.9em;">Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}</p>
        </div>
        
        <div class="stats">
            <div class="stat">
                <div class="stat-value">{len(fixes)}</div>
                <div class="stat-label">Total Fixes</div>
            </div>
            <div class="stat">
                <div class="stat-value">{sum(1 for f in fixes if f['fix_suggestion']['type'] == 'parasoft_official')}</div>
                <div class="stat-label">Parasoft Official</div>
            </div>
            <div class="stat">
                <div class="stat-value">{sum(1 for f in fixes if f['fix_suggestion'].get('priority') in ['CRITICAL', 'HIGH'])}</div>
                <div class="stat-label">High Priority</div>
            </div>
            <div class="stat">
                <div class="stat-value">{len(set(f['category'] for f in fixes))}</div>
                <div class="stat-label">Categories</div>
            </div>
        </div>
        
        <div class="filters">
            <div class="filter-group">
                <label>Search:</label>
                <input type="text" id="searchInput" placeholder="Filter by violation ID or description..." style="width: 300px;" onkeyup="filterFixes()">
            </div>
            <div class="filter-group">
                <label>Priority:</label>
                <select id="priorityFilter" onchange="filterFixes()">
                    <option value="">All</option>
                    <option value="CRITICAL">Critical</option>
                    <option value="HIGH">High</option>
                    <option value="MEDIUM">Medium</option>
                    <option value="LOW">Low</option>
                </select>
            </div>
            <div class="filter-group">
                <label>Type:</label>
                <select id="typeFilter" onchange="filterFixes()">
                    <option value="">All</option>
                    <option value="parasoft_official">Parasoft Official</option>
                    <option value="ai_generated">AI Generated</option>
                    <option value="rule_based">Rule Based</option>
                </select>
            </div>
        </div>
        
        <div class="fixes-container" id="fixesContainer">
'''
        
        # Generate fix cards
        for i, fix in enumerate(fixes, 1):
            fix_suggestion = fix['fix_suggestion']
            priority = fix_suggestion.get('priority', 'MEDIUM')
            fix_type = fix_suggestion['type']
            
            # Check if this is a justification recommendation
            is_justification_rec = (fix_type == 'justification_recommended')
            
            # Determine badge colors
            priority_class = f"badge-{priority.lower()}"
            type_display = {
                'parasoft_official': ('Parasoft Official', 'badge-parasoft'),
                'ai_generated': ('AI Generated', 'badge-ai'),
                'rule_based': ('Rule Based', 'badge-rule'),
                'justification_recommended': ('⚖️ Justify Instead', 'badge-info')
            }.get(fix_type, (fix_type.title(), 'badge-rule'))
            
            html_content += f'''
            <div class="fix-card {'justification-rec-card' if is_justification_rec else ''}" data-priority="{priority}" data-type="{fix_type}" data-violation="{fix['violation_id']}" data-description="{fix['violation_text']}">
                <div class="fix-header">
                    <div class="fix-title">{'⚖️  JUSTIFICATION SUGGESTION' if is_justification_rec else 'FIX'} #{i} - {fix['violation_id']}</div>
                    <div class="fix-badges">
                        <span class="badge {priority_class}">{priority}</span>
                        <span class="badge {type_display[1]}">{type_display[0]}</span>
                        <span class="badge">{fix['severity']}</span>
                    </div>
                </div>
                <div class="fix-content">
                    <div class="section">
                        <div class="section-title">📋 Violation Description</div>
                        <div class="description">{fix['violation_text']}</div>
                    </div>
                    
                    <div class="section">
                        <div class="section-title">{'⚖️  AI Recommendation' if is_justification_rec else '💡 Fix Suggestion'}</div>
                        <div class="description">{fix_suggestion.get('description', 'No description available')}</div>
                    </div>
'''
            
            # Add code examples if available
            if 'example' in fix_suggestion and fix_suggestion['example']:
                example = fix_suggestion['example']
                
                # Check if it's Parasoft format with special markers
                if 'VIOLATION EXAMPLE' in example or 'OFFICIAL REPAIR' in example:
                    if 'VIOLATION EXAMPLE' in example:
                        violation_part = example.split('--- OFFICIAL REPAIR ---')[0].replace('--- VIOLATION EXAMPLE ---', '').strip()
                        if violation_part:
                            html_content += f'''
                    <div class="section">
                        <div class="code-header violation-code">❌ Violation Example</div>
                        <div class="code-block"><pre>{violation_part}</pre></div>
                    </div>
'''
                    
                    if 'OFFICIAL REPAIR' in example:
                        repair_part = example.split('--- OFFICIAL REPAIR ---')[1].strip() if '--- OFFICIAL REPAIR ---' in example else ''
                        if repair_part:
                            html_content += f'''
                    <div class="section">
                        <div class="code-header repair-code">✅ Official Repair</div>
                        <div class="code-block"><pre>{repair_part}</pre></div>
                    </div>
'''
                else:
                    # Regular AI-generated or pattern-based example
                    # Display as single code block with before/after
                    html_content += f'''
                    <div class="section">
                        <div class="code-header">📝 Code Example</div>
                        <div class="code-block"><pre>{example}</pre></div>
                    </div>
'''
            
            # Add CWE mappings if available
            if 'cwe_mappings' in fix_suggestion and fix_suggestion['cwe_mappings']:
                html_content += f'''
                    <div class="section">
                        <div class="section-title">🔒 Security Mappings</div>
                        <div class="cwe-tags">
'''
                for cwe in fix_suggestion['cwe_mappings'][:8]:  # Limit to 8
                    html_content += f'                            <span class="cwe-tag">{cwe}</span>\n'
                html_content += '''                        </div>
                    </div>
'''
            
            # Add affected files
            files_to_show = fix['files_affected'][:10]  # Show first 10
            html_content += f'''
                    <div class="section">
                        <div class="section-title">📁 Affected Files ({len(fix['files_affected'])} total)</div>
                        <div class="files-list">
'''
            for file_entry in files_to_show:
                html_content += f'                            <div class="file-item">{file_entry["file"]}:{file_entry["line"]}</div>\n'
            
            if len(fix['files_affected']) > 10:
                html_content += f'                            <div class="file-item" style="color: #666; font-style: italic;">... and {len(fix["files_affected"]) - 10} more files</div>\n'
            
            html_content += '''                        </div>
                    </div>
'''
            
            # Add source info if Parasoft official
            if 'source' in fix_suggestion:
                html_content += f'''
                    <div class="section">
                        <div class="section-title">📚 Source</div>
                        <div class="description" style="font-style: italic;">{fix_suggestion['source']}</div>
                    </div>
'''
            
            html_content += '''                </div>
            </div>
'''
        
        # Footer and scripts
        html_content += f'''        </div>
        
        <div class="footer">
            <div class="footer-logo">
                <svg viewBox="0 0 200 200" xmlns="http://www.w3.org/2000/svg">
                    <circle cx="100" cy="100" r="90" fill="#667eea" opacity="0.2"/>
                    <text x="100" y="120" font-size="80" font-weight="bold" text-anchor="middle" fill="#667eea">Q</text>
                </svg>
            </div>
            <p style="font-weight: 600; margin-bottom: 10px;">Qorix India Pvt Ltd</p>
            <p>Parasoft Analysis Tool v2.3.0 - Powered by Official Parasoft Rules Database</p>
            <p style="margin-top: 10px; font-size: 0.9em;">© 2025 Qorix India Pvt Ltd. All Rights Reserved.</p>
        </div>
    </div>
    
    <script>
        function filterFixes() {{
            const searchTerm = document.getElementById('searchInput').value.toLowerCase();
            const priorityFilter = document.getElementById('priorityFilter').value;
            const typeFilter = document.getElementById('typeFilter').value;
            
            const fixCards = document.querySelectorAll('.fix-card');
            let visibleCount = 0;
            
            fixCards.forEach(card => {{
                const violationId = card.getAttribute('data-violation').toLowerCase();
                const description = card.getAttribute('data-description').toLowerCase();
                const priority = card.getAttribute('data-priority');
                const type = card.getAttribute('data-type');
                
                const matchesSearch = !searchTerm || violationId.includes(searchTerm) || description.includes(searchTerm);
                const matchesPriority = !priorityFilter || priority === priorityFilter;
                const matchesType = !typeFilter || type === typeFilter;
                
                if (matchesSearch && matchesPriority && matchesType) {{
                    card.style.display = 'block';
                    visibleCount++;
                }} else {{
                    card.style.display = 'none';
                }}
            }});
            
            // Update stats if needed
            console.log(`Showing ${{visibleCount}} of ${{fixCards.length}} fixes`);
        }}
        
        // Make fix headers collapsible
        document.querySelectorAll('.fix-header').forEach(header => {{
            header.addEventListener('click', function() {{
                const content = this.nextElementSibling;
                if (content.style.display === 'none') {{
                    content.style.display = 'block';
                }} else {{
                    content.style.display = 'none';
                }}
            }});
        }});
    </script>
</body>
</html>
'''
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"HTML report saved to: {output_file}")
    
    def _save_fixes_file(self, fixes: List[Dict], output_file: Path):
        """
        Save fixes to a text file
        
        Args:
            fixes: List of fix dictionaries
            output_file: Path to output file
        """
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write(f"PARASOFT FIXES - Module: {self.module_name}\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total Fixes: {len(fixes)}\n")
            f.write("="*80 + "\n\n")
            
            for i, fix in enumerate(fixes, 1):
                fix_type = fix['fix_suggestion']['type']
                is_justification_rec = (fix_type == 'justification_recommended')
                
                f.write(f"{'⚖️  JUSTIFICATION SUGGESTION' if is_justification_rec else 'FIX'} #{i}\n")
                f.write("-"*80 + "\n")
                f.write(f"Violation ID: {fix['violation_id']}\n")
                f.write(f"Severity: {fix['severity']}\n")
                f.write(f"Category: {fix['category']}\n")
                if is_justification_rec:
                    f.write(f"🔍 Recommended Action: JUSTIFY (Add to common deviations Excel)\n")
                f.write("\n")
                
                f.write(f"Description:\n{fix['violation_text']}\n\n")
                
                f.write(f"{'Recommendation' if is_justification_rec else 'Fix'} Type: {fix_type}\n")
                f.write(f"Priority: {fix['fix_suggestion']['priority']}\n\n")
                
                f.write(f"{'AI Analysis' if is_justification_rec else 'Fix Description'}:\n{fix['fix_suggestion']['description']}\n\n")
                
                if 'example' in fix['fix_suggestion']:
                    if is_justification_rec:
                        f.write(f"Cross-Module Context & Suggested Rationale:\n{fix['fix_suggestion']['example']}\n\n")
                    else:
                        f.write(f"Example:\n{fix['fix_suggestion']['example']}\n\n")
                
                f.write(f"Files Affected:\n")
                files = fix.get('files_affected', [])
                if files:
                    for file_entry in files[:5]:
                        # Handle both string and dict formats
                        if isinstance(file_entry, dict):
                            f.write(f"  - {file_entry.get('file', 'unknown')}:{file_entry.get('line', '?')}\n")
                        else:
                            f.write(f"  - {file_entry}\n")
                    
                    if len(files) > 5:
                        f.write(f"  ... and {len(files) - 5} more files\n")
                else:
                    f.write(f"  - No file information available\n")
                
                f.write("\n" + "="*80 + "\n\n")
    
    def _add_justification_sheet_to_excel(self, justification_recommendations: List[Dict]) -> bool:
        """
        Add a new sheet to violations Excel report with AI justification recommendations
        
        Args:
            justification_recommendations: List of violation dicts with justification recommendations
        
        Returns:
            True if successfully added, False otherwise
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available - cannot update Excel report")
            return False
        
        try:
            # Find the violations report Excel file
            report_file = self._find_violations_report_excel()
            if not report_file:
                logger.warning("Violations report Excel file not found")
                return False
            
            # Load the workbook
            workbook = load_workbook(report_file)
            
            # Remove sheet if it already exists (update scenario)
            sheet_name = "AI Justification Suggestions"
            if sheet_name in workbook.sheetnames:
                logger.info(f"Removing existing '{sheet_name}' sheet")
                del workbook[sheet_name]
            
            # Create new sheet
            ws = workbook.create_sheet(sheet_name)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Headers
            headers = [
                "Violation ID",
                "Category",
                "Severity", 
                "Description",
                "AI Confidence",
                "Suggested Rationale",
                "Cross-Module Count",
                "Modules with Justifications",
                "Files Affected Count",
                "Recommendation"
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Write data rows
            for row_idx, rec in enumerate(justification_recommendations, 2):
                justification_analysis = rec.get('fix_suggestion', {}).get('justification_analysis', {})
                
                # Extract cross-module info from example field
                example_text = rec.get('fix_suggestion', {}).get('example', '')
                cross_module_count = 0
                modules_list = []
                
                # Parse cross-module info from example text
                if 'Modules with justifications:' in example_text:
                    lines = example_text.split('\n')
                    for line in lines:
                        if line.strip().startswith('- Module'):
                            modules_list.append(line.strip().replace('- ', ''))
                    cross_module_count = len(modules_list)
                
                # Extract suggested rationale
                suggested_rationale = justification_analysis.get('suggested_rationale', '')
                if not suggested_rationale and 'Reason:' in example_text:
                    for line in example_text.split('\n'):
                        if line.startswith('Reason:'):
                            suggested_rationale = line.replace('Reason:', '').strip()
                            break
                
                # Row data
                row_data = [
                    rec.get('violation_id', ''),
                    rec.get('category', ''),
                    rec.get('severity', ''),
                    rec.get('violation_text', ''),
                    justification_analysis.get('confidence', 'MEDIUM'),
                    suggested_rationale,
                    cross_module_count,
                    '\n'.join(modules_list),
                    len(rec.get('files_affected', [])),
                    "Add to Common Deviations Excel"
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = border
                    
                    # Highlight high confidence recommendations
                    if col == 5 and value == 'HIGH':
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        cell.font = Font(bold=True, color="155724")
            
            # Adjust column widths
            column_widths = [20, 12, 10, 50, 12, 40, 12, 40, 12, 30]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Add summary at top (insert rows)
            ws.insert_rows(1, 3)
            ws.merge_cells('A1:J1')
            ws['A1'] = f"AI Justification Recommendations - {self.module_name}"
            ws['A1'].font = Font(bold=True, size=14, color="4472C4")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            
            ws.merge_cells('A2:J2')
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Recommendations: {len(justification_recommendations)}"
            ws['A2'].font = Font(italic=True, size=10)
            ws['A2'].alignment = Alignment(horizontal="center")
            
            ws.merge_cells('A3:J3')
            ws['A3'] = "⚠️ These violations appear in multiple modules as justified. Consider adding to common deviations Excel instead of fixing."
            ws['A3'].font = Font(bold=True, size=10, color="856404")
            ws['A3'].fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            ws['A3'].alignment = Alignment(horizontal="center", wrap_text=True)
            
            # Save workbook
            workbook.save(report_file)
            logger.info(f"Updated Excel report: {report_file}")
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to add justification sheet to Excel: {str(e)}")
            return False
    
    def _find_violations_report_excel(self) -> Optional[Path]:
        """
        Find the violations report Excel file
        
        Returns:
            Path to violations report or None
        """
        # Expected filename pattern
        report_filename = f"{self.module_name}_violations_report.xlsx"
        
        # Search in common locations
        search_paths = [
            self.reports_dir / report_filename,  # reports/ folder
            Path.cwd() / "reports" / report_filename,
            Path.cwd() / report_filename,
            self.reports_dir.parent / report_filename,  # Parent of reports
        ]
        
        for path in search_paths:
            if path.exists():
                logger.debug(f"Found violations report: {path}")
                return path
        
        logger.debug(f"Violations report not found. Searched: {[str(p) for p in search_paths]}")
        return None
    
    def generate_justifications(self, violation_ids: Optional[List[str]] = None) -> Dict:
        """
        Generate Parasoft-formatted justifications
        
        Args:
            violation_ids: List of specific violation IDs (None = all unjustified)
        
        Returns:
            Results dictionary
        """
        if violation_ids:
            violations_to_justify = [
                self.kb_manager.get_violation(vid)
                for vid in violation_ids
                if self.kb_manager.get_violation(vid)
            ]
        else:
            violations_to_justify = [
                v for v in self.kb_manager.get_all_violations()
                if not v.get('justification_added') and v.get('justifiable') == 'Yes'
            ]
        
        logger.info(f"Generating justifications for {len(violations_to_justify)} violations")
        
        justifications_generated = 0
        all_justifications = []
        
        for violation in violations_to_justify:
            just_data = self._generate_justification(violation)
            if just_data:
                all_justifications.append(just_data)
                justifications_generated += 1
        
        # Save justifications to file
        if all_justifications:
            just_file = self.module_fixes_dir / f"{self.module_name}_justifications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            self._save_justifications_file(all_justifications, just_file)
            logger.info(f"Justifications saved to: {just_file}")
        
        return {
            'justifications_generated': justifications_generated,
            'total_violations': len(violations_to_justify),
            'justifications_file': str(just_file) if all_justifications else None
        }
    
    def _generate_justification(self, violation: Dict) -> Optional[Dict]:
        """
        Generate Parasoft-formatted justification
        
        Args:
            violation: Violation dictionary
        
        Returns:
            Justification data dictionary
        """
        violation_id = violation['violation_id']
        
        # Generate justification text in Parasoft format
        justification_text = self._format_parasoft_justification(violation)
        
        just_data = {
            'violation_id': violation_id,
            'violation_text': violation['violation_text'],
            'justification': justification_text,
            'files_affected': violation['files_affected'],
            'timestamp': datetime.now().isoformat()
        }
        
        # Update knowledge base
        self.kb_manager.update_justification_status(violation_id, justification_text)
        
        return just_data
    
    def _format_parasoft_justification(self, violation: Dict) -> str:
        """
        Format justification in Parasoft format
        
        Args:
            violation: Violation dictionary
        
        Returns:
            Formatted justification text
        """
        # Parasoft justification format:
        # /* parasoft-begin-suppress RULE_ID "Reason for suppression" */
        # ... code ...
        # /* parasoft-end-suppress RULE_ID */
        
        violation_id = violation['violation_id']
        
        # Generate reason based on violation type
        reason = self._generate_justification_reason(violation)
        
        justification = f'''/* parasoft-begin-suppress {violation_id} "{reason}" */
/* 
 * Justification: {reason}
 * Reviewed by: [Developer Name]
 * Date: {datetime.now().strftime('%Y-%m-%d')}
 * Ticket/Issue: [Reference]
 */
... your code here ...
/* parasoft-end-suppress {violation_id} */'''
        
        return justification
    
    def _generate_justification_reason(self, violation: Dict) -> str:
        """
        Generate justification reason using Parasoft Rules Database (PRIORITY) -> Fallback to generic
        
        Args:
            violation: Violation dictionary
        
        Returns:
            Justification reason text
        """
        violation_id = violation['violation_id']
        violation_text = violation['violation_text'].upper()
        
        # PRIORITY: Try to get official rationale from Parasoft Rules Database
        if self.use_rules_db and self.rules_parser:
            official_reason = self._get_parasoft_justification_rationale(violation_id)
            if official_reason:
                # Use official security relevance + benefits as justification reason
                return f"Parasoft Official: {official_reason[:200]}"
        
        # FALLBACK: Common generic justification reasons
        if 'DESIGN' in violation_text or 'ARCHITECTURE' in violation_text:
            return "Design decision - alternative approach not feasible in current architecture"
        
        elif 'LEGACY' in violation_text or 'THIRD PARTY' in violation_text:
            return "Legacy code compatibility - cannot be changed without breaking existing functionality"
        
        elif 'PERFORMANCE' in violation_text:
            return "Performance optimization - critical path where alternative approach impacts performance"
        
        elif 'STANDARD LIBRARY' in violation_text:
            return "Required for platform compatibility - no safer alternative available"
        
        else:
            return "Technical constraint - fix would require significant architectural changes"
    
    def _get_parasoft_justification_rationale(self, violation_id: str) -> Optional[str]:
        """
        Get official justification rationale from Parasoft Rules Database
        
        Args:
            violation_id: Violation ID
        
        Returns:
            Justification rationale text combining security relevance and benefits, or None
        """
        try:
            # Get rule from database
            rationale = self.rules_parser.get_justification_rationale(violation_id)
            if rationale:
                # Clean up and format
                rationale = rationale.replace('\n', ' ').strip()
                # Limit length
                if len(rationale) > 250:
                    rationale = rationale[:247] + "..."
                return rationale
            return None
        except Exception as e:
            logger.debug(f"Could not extract justification rationale for {violation_id}: {e}")
            return None
    
    def _save_justifications_file(self, justifications: List[Dict], output_file: Path):
        """
        Save justifications to a text file
        
        Args:
            justifications: List of justification dictionaries
            output_file: Path to output file
        """
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write(f"PARASOFT JUSTIFICATIONS - Module: {self.module_name}\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total Justifications: {len(justifications)}\n")
            f.write("="*80 + "\n\n")
            
            for i, just in enumerate(justifications, 1):
                f.write(f"JUSTIFICATION #{i}\n")
                f.write("-"*80 + "\n")
                f.write(f"Violation ID: {just['violation_id']}\n\n")
                
                f.write(f"Description:\n{just['violation_text']}\n\n")
                
                f.write(f"Parasoft Justification Format:\n")
                f.write(just['justification'] + "\n\n")
                
                f.write(f"Apply to files:\n")
                for file_entry in just['files_affected'][:5]:
                    f.write(f"  - {file_entry['file']}:{file_entry['line']}\n")
                
                if len(just['files_affected']) > 5:
                    f.write(f"  ... and {len(just['files_affected']) - 5} more files\n")
                
                f.write("\n" + "="*80 + "\n\n")
